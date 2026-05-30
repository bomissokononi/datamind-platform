"""
╔══════════════════════════════════════════════════════════════╗
║         DataMind AI Platform v3.0 — Fichier unique          ║
║  Big Data Analytics × IA Auto-Interrogative × Rapport PDF   ║
╚══════════════════════════════════════════════════════════════╝
"""
import os, json, time, warnings, concurrent.futures, io
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import httpx
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy import stats
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from dotenv import load_dotenv

warnings.filterwarnings("ignore")
load_dotenv()

OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Modèles gratuits OpenRouter (liste vérifiée mai 2025) ────────────────────
# ── Modèles gratuits OpenRouter — vérifiés mai 2026 ─────────────────────────
FREE_MODELS = [
    # Routeur automatique — choisit le meilleur modèle gratuit disponible
    "openrouter/auto",                              # ✅ Routeur auto OpenRouter
    # LLaMA (Meta) — très stables
    "meta-llama/llama-3.3-70b-instruct:free",       # ✅ Meilleur général
    "meta-llama/llama-3.1-8b-instruct:free",        # ✅ Rapide
    # DeepSeek — excellent raisonnement
    "deepseek/deepseek-r1-0528:free",               # ✅ Raisonnement avancé
    "deepseek/deepseek-v3-0324:free",               # ✅ Analyse générale
    # Qwen (Alibaba) — excellent multilingue/français
    "qwen/qwen3-14b:free",                          # ✅ Multilingue
    "qwen/qwen3-8b:free",                           # ✅ Léger
    "qwen/qwen3-235b-a22b:free",                    # ✅ Très puissant
    # Google
    "google/gemma-3-27b-it:free",                   # ✅ Bon général
    "google/gemma-3n-e4b-it:free",                  # ✅ Compact
    # Microsoft
    "microsoft/phi-4-reasoning-plus:free",          # ✅ Raisonnement
    # NVIDIA
    "nvidia/llama-3.3-nemotron-super-49b-v1:free",  # ✅ Puissant
    # OpenAI open source
    "openai/gpt-4o-mini-search-preview:free",       # ✅ GPT avec recherche
    # Mistral
    "mistralai/mistral-7b-instruct:free",           # ✅ Bon en français
    # Autres
    "thudm/glm-4-32b:free",                         # ✅ Multilingue
    "moonshotai/moonlight-16b-a3b-instruct:free",   # ✅ Efficace
]

# Modèles recommandés par usage
RECOMMENDED_MODELS = {
    "analyse":     "meta-llama/llama-3.3-70b-instruct:free",
    "raisonnement":"deepseek/deepseek-r1-0528:free",
    "français":    "qwen/qwen3-14b:free",
    "rapide":      "meta-llama/llama-3.1-8b-instruct:free",
    "auto":        "openrouter/auto",
}

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

# ══════════════════════════════════════════════════════════════════════════════
# 1. PIPELINE DE DONNÉES
# ══════════════════════════════════════════════════════════════════════════════

def auto_clean(df):
    report = {}
    n_dup = int(df.duplicated().sum())
    df = df.drop_duplicates()
    report["doublons_supprimés"] = n_dup
    missing = {}
    for col in df.columns:
        pct = df[col].isna().mean()
        if pct > 0.8:
            df = df.drop(columns=[col])
            missing[col] = f"Colonne supprimée (>{80}% manquants)"
        elif pct > 0:
            if df[col].dtype in [np.float64, np.int64, float, int]:
                df[col] = df[col].fillna(df[col].median())
            else:
                mode = df[col].mode()
                df[col] = df[col].fillna(mode.iloc[0] if not mode.empty else "INCONNU")
            missing[col] = f"{pct:.1%} valeurs remplies"
    report["valeurs_manquantes"] = missing
    df.columns = (df.columns.str.strip().str.lower()
                  .str.replace(r"[^\w]", "_", regex=True)
                  .str.replace(r"_+", "_", regex=True).str.strip("_"))
    for col in df.columns:
        if df[col].dtype == object:
            try:
                df[col] = pd.to_numeric(df[col], errors="raise")
            except:
                try:
                    df[col] = pd.to_datetime(df[col], infer_datetime_format=True, errors="raise")
                except:
                    pass
    return df, report


def full_eda(df):
    eda = {}
    num_cols = df.select_dtypes(include=np.number).columns.tolist()
    cat_cols = df.select_dtypes(include=["object","category"]).columns.tolist()
    dt_cols  = df.select_dtypes(include=["datetime64"]).columns.tolist()

    eda["overview"] = {
        "rows": int(df.shape[0]), "columns": int(df.shape[1]),
        "numeric_cols": num_cols, "categorical_cols": cat_cols, "datetime_cols": dt_cols,
        "memory_mb": round(df.memory_usage(deep=True).sum()/1e6,2),
        "missing_pct": round(df.isna().mean().mean()*100,2),
    }

    if num_cols:
        desc = df[num_cols].describe(percentiles=[.1,.25,.5,.75,.9]).round(4)
        eda["descriptive_stats"] = desc.to_dict()

    eda["distributions"] = {}
    for col in num_cols:
        s = df[col].dropna()
        if len(s) > 3:
            sk = float(stats.skew(s))
            ku = float(stats.kurtosis(s))
            try:
                _, pval = stats.shapiro(s.sample(min(5000,len(s)), random_state=42))
            except:
                pval = 0
            eda["distributions"][col] = {
                "skewness": round(sk,3), "kurtosis": round(ku,3),
                "normality_pvalue": round(float(pval),4),
                "is_normal": pval > 0.05,
                "q1": round(float(s.quantile(.25)),4),
                "q3": round(float(s.quantile(.75)),4),
                "iqr": round(float(s.quantile(.75)-s.quantile(.25)),4),
                "outliers_iqr": int(((s < s.quantile(.25)-1.5*(s.quantile(.75)-s.quantile(.25)))|(s > s.quantile(.75)+1.5*(s.quantile(.75)-s.quantile(.25)))).sum()),
            }

    if len(num_cols) > 1:
        corr = df[num_cols].corr()
        strong = []
        for i in range(len(corr.columns)):
            for j in range(i+1, len(corr.columns)):
                v = corr.iloc[i,j]
                if abs(v) > 0.4:
                    strong.append({"col1":corr.columns[i],"col2":corr.columns[j],"r":round(float(v),3)})
        eda["correlations"] = {
            "matrix": corr.round(3).to_dict(),
            "strong_pairs": sorted(strong, key=lambda x: abs(x["r"]), reverse=True),
        }

    eda["categorical_analysis"] = {}
    for col in cat_cols:
        vc = df[col].value_counts()
        total = len(df[col].dropna())
        eda["categorical_analysis"][col] = {
            "unique_count": int(df[col].nunique()),
            "top_values": vc.head(10).to_dict(),
            "top_pct": {k: round(v/total*100,1) for k,v in vc.head(10).items()},
        }

    return eda


def detect_anomalies(df, contamination=0.05):
    num_cols = df.select_dtypes(include=np.number).columns.tolist()
    if not num_cols:
        df["anomalie"] = False
        return df
    X = df[num_cols].fillna(df[num_cols].median())
    model = IsolationForest(contamination=contamination, random_state=42, n_jobs=-1)
    df = df.copy()
    preds = model.fit_predict(X)
    df["anomalie"] = preds == -1
    df["score_anomalie"] = model.decision_function(X).round(4)
    return df


def compute_pca(df, n_components=2):
    num_cols = df.select_dtypes(include=np.number).columns.tolist()
    cols = [c for c in num_cols if c not in ["anomalie","score_anomalie"]]
    if len(cols) < 2:
        return None, None
    X = df[cols].fillna(df[cols].median())
    scaler = StandardScaler()
    Xs = scaler.fit_transform(X)
    n = min(n_components, len(cols))
    pca = PCA(n_components=n, random_state=42)
    Xp = pca.fit_transform(Xs)
    explained = pca.explained_variance_ratio_
    df_pca = pd.DataFrame(Xp, columns=[f"PC{i+1}" for i in range(n)])
    return df_pca, explained


# ══════════════════════════════════════════════════════════════════════════════
# 2. MOTEUR IA OPENROUTER
# ══════════════════════════════════════════════════════════════════════════════

def call_openrouter(messages, model, api_key, system="", max_tokens=3000, temperature=0.3, retries=3):
    if not api_key:
        return "❌ Clé API manquante — ajoutez OPENROUTER_API_KEY dans la barre latérale."
    # openrouter/auto et certains modèles n'ont pas le suffixe :free
    if not model.endswith(":free") and not model.startswith("openrouter/"):
        model += ":free"
    msgs = ([{"role":"system","content":system}] if system else []) + messages
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://datamind.ai",
        "X-Title": "DataMind Platform",
    }
    last_error = ""
    for attempt in range(1, retries+1):
        try:
            r = httpx.post(OPENROUTER_URL, headers=headers,
                           json={"model":model,"messages":msgs,"max_tokens":max_tokens,"temperature":temperature},
                           timeout=120)
            if r.status_code == 429:
                wait = 15 * attempt
                time.sleep(wait)
                last_error = f"Rate limit (429) — attente {wait}s"
                continue
            if r.status_code == 404:
                last_error = f"Modèle introuvable (404) : {model}"
                break
            if r.status_code == 401:
                return "❌ Clé API invalide (401). Vérifiez votre clé sur openrouter.ai"
            if r.status_code == 402:
                return "❌ Crédits insuffisants (402). Rechargez votre compte openrouter.ai"
            if r.status_code != 200:
                try:
                    err_detail = r.json().get("error", {}).get("message", r.text[:300])
                except:
                    err_detail = r.text[:300]
                last_error = f"Erreur HTTP {r.status_code} : {err_detail}"
                if attempt < retries:
                    time.sleep(5 * attempt)
                continue
            data = r.json()
            if "choices" not in data or not data["choices"]:
                last_error = f"Réponse vide du modèle {model}"
                continue
            text = data["choices"][0]["message"]["content"]
            if not text or not text.strip():
                last_error = f"Contenu vide reçu de {model}"
                continue
            return text
        except httpx.TimeoutException:
            last_error = f"Timeout (120s) — tentative {attempt}/{retries}"
            if attempt < retries:
                time.sleep(5 * attempt)
        except httpx.ConnectError:
            last_error = "Impossible de se connecter à OpenRouter. Vérifiez votre connexion internet."
            break
        except Exception as e:
            last_error = f"Erreur inattendue : {type(e).__name__}: {e}"
            break
    return f"❌ {last_error}"


def call_with_fallback(messages, preferred_model, api_key, system="", max_tokens=3000, temperature=0.3):
    """Essaie le modèle préféré, puis bascule automatiquement sur les suivants."""
    # Ordre de fallback : modèle choisi → auto → llama 70B → qwen → deepseek → llama 8B
    fallback_order = [
        preferred_model,
        "openrouter/auto",
        "meta-llama/llama-3.3-70b-instruct:free",
        "qwen/qwen3-14b:free",
        "deepseek/deepseek-v3-0324:free",
        "meta-llama/llama-3.1-8b-instruct:free",
    ]
    # Dédupliquer en gardant l'ordre
    seen = set()
    all_models = []
    for m in fallback_order:
        if m not in seen:
            seen.add(m)
            all_models.append(m)

    last_error = "Aucune tentative effectuée"
    for model in all_models:
        result = call_openrouter(messages, model, api_key, system=system,
                                  max_tokens=max_tokens, temperature=temperature, retries=2)
        if not result.startswith("❌"):
            return result, model
        last_error = result

    return f"❌ Tous les modèles ont échoué. Vérifiez votre clé API sur openrouter.ai\nDernière erreur : {last_error}", preferred_model


def generate_ai_questions(eda, context, language, model, api_key):
    """L'IA génère elle-même des questions pertinentes sur les données."""
    ov = eda.get("overview", {})
    lang = "français" if language == "fr" else "English"
    prompt = f"""Tu es un data scientist expert. Voici les métadonnées d'un jeu de données :
- {ov.get('rows','?')} lignes × {ov.get('columns','?')} colonnes
- Variables numériques : {', '.join(ov.get('numeric_cols',[])[:12])}
- Variables catégorielles : {', '.join(ov.get('categorical_cols',[])[:8])}
- Variables temporelles : {', '.join(ov.get('datetime_cols',[])[:4])}
- Contexte : {context or 'Non précisé'}

Génère 10 questions analytiques PERTINENTES et PROFONDES que l'on devrait se poser sur ces données.
Chaque question doit :
1. Être spécifique aux colonnes présentes
2. Apporter une valeur business réelle
3. Être accompagnée d'une hypothèse de réponse

Réponds en {lang} avec ce format exact :
Q1: [question]
Hypothèse: [hypothèse de réponse attendue]

Q2: [question]
Hypothèse: [hypothèse]
... etc."""

    system = f"Tu es DataMind, expert en analyse de données. Réponds en {lang}."
    result, used_model = call_with_fallback([{"role":"user","content":prompt}], model, api_key, system=system, max_tokens=2000)
    return result


def answer_ai_questions(eda, questions_text, context, language, model, api_key):
    """L'IA répond à ses propres questions avec les données disponibles."""
    ov = eda.get("overview",{})
    ds = eda.get("descriptive_stats",{})
    corr = eda.get("correlations",{}).get("strong_pairs",[])
    dist = eda.get("distributions",{})
    lang = "français" if language == "fr" else "English"

    stats_summary = []
    for col, s in list(ds.items())[:10]:
        try:
            stats_summary.append(f"{col}: moy={float(s.get('mean',0)):.3g}, σ={float(s.get('std',0)):.3g}, min={float(s.get('min',0)):.3g}, max={float(s.get('max',0)):.3g}")
        except: pass

    outliers_summary = []
    for col, d in list(dist.items())[:8]:
        if d.get("outliers_iqr",0) > 0:
            outliers_summary.append(f"{col}: {d['outliers_iqr']} outliers, skewness={d['skewness']}")

    prompt = f"""Tu es DataMind, analyste expert. Voici les données statistiques réelles :

STATISTIQUES :
{chr(10).join(stats_summary)}

CORRÉLATIONS FORTES :
{chr(10).join([f"{p['col1']} ↔ {p['col2']}: r={p['r']}" for p in corr[:6]])}

OUTLIERS DÉTECTÉS :
{chr(10).join(outliers_summary) or 'Aucun notable'}

CONTEXTE MÉTIER : {context or 'Non précisé'}

QUESTIONS À RÉPONDRE :
{questions_text}

Pour chaque question, donne une réponse analytique basée sur les données ci-dessus.
Cite des chiffres précis. Indique le niveau de certitude. Formule des recommandations.
Réponds en {lang}."""

    system = f"Tu es un analyste de données senior. Réponds en {lang} avec précision et rigueur."
    result, used_model = call_with_fallback([{"role":"user","content":prompt}], model, api_key, system=system, max_tokens=3000)
    return result


def build_deep_analysis_prompt(eda, context, language):
    ov = eda.get("overview",{})
    ds = eda.get("descriptive_stats",{})
    corr = eda.get("correlations",{}).get("strong_pairs",[])
    dist = eda.get("distributions",{})
    cat  = eda.get("categorical_analysis",{})
    lang = "français" if language == "fr" else "English"

    lines = [
        f"═══ DONNÉES : {ov.get('rows','?')} lignes × {ov.get('columns','?')} colonnes ═══",
        f"Mémoire : {ov.get('memory_mb','?')} Mo | Manquants : {ov.get('missing_pct','?')}%",
        f"Numériques    : {', '.join(ov.get('numeric_cols',[])[:12])}",
        f"Catégorielles : {', '.join(ov.get('categorical_cols',[])[:8])}",
        f"Temporelles   : {', '.join(ov.get('datetime_cols',[])[:4])}",
    ]
    if ds:
        lines.append("\n── STATISTIQUES DESCRIPTIVES ──")
        for col, s in list(ds.items())[:10]:
            try:
                lines.append(f"  {col}: moy={float(s.get('mean',0)):.4g}, σ={float(s.get('std',0)):.4g}, "
                             f"min={float(s.get('min',0)):.4g}, max={float(s.get('max',0)):.4g}, "
                             f"médiane={float(s.get('50%',0)):.4g}")
            except: pass
    if corr:
        lines.append("\n── CORRÉLATIONS ──")
        for p in corr[:8]:
            force = "très forte" if abs(p['r'])>0.8 else "forte" if abs(p['r'])>0.65 else "modérée"
            dir_ = "positive" if p['r']>0 else "négative"
            lines.append(f"  {p['col1']} ↔ {p['col2']}: r={p['r']} ({force}, {dir_})")
    if dist:
        lines.append("\n── DISTRIBUTIONS ──")
        for col, d in list(dist.items())[:8]:
            norm = "normale" if d.get("is_normal") else "non-normale"
            lines.append(f"  {col}: skew={d['skewness']}, kurt={d['kurtosis']}, "
                        f"outliers={d['outliers_iqr']}, distribution={norm}")
    if cat:
        lines.append("\n── CATÉGORIELLES ──")
        for col, info in list(cat.items())[:5]:
            top = list(info.get("top_values",{}).keys())[:3]
            pcts = list(info.get("top_pct",{}).values())[:3]
            lines.append(f"  {col} ({info['unique_count']} valeurs): Top → {', '.join(f'{v}({p}%)' for v,p in zip(top,pcts))}")
    if context:
        lines.append(f"\n── CONTEXTE MÉTIER ──\n{context}")

    lines.append(f"""
═══ MISSION ANALYTIQUE ═══
Génère un rapport d'analyse COMPLET et DÉTAILLÉ en {lang} avec :

## 1. RÉSUMÉ EXÉCUTIF
   - Vue d'ensemble des données (qualité, complétude, pertinence)
   - 3-5 phrases synthétisant les découvertes majeures

## 2. ANALYSE STATISTIQUE APPROFONDIE
   - Interprétation des distributions (normalité, asymétrie, extremes)
   - Analyse des valeurs aberrantes et leur impact
   - Tendances centrales et dispersions remarquables

## 3. RELATIONS ET CORRÉLATIONS
   - Explication des corrélations fortes avec contexte métier
   - Identification de causalités potentielles
   - Variables redondantes ou complémentaires

## 4. SEGMENTATION ET PATTERNS
   - Groupes naturels identifiables dans les données
   - Comportements récurrents ou cycliques
   - Anomalies structurelles

## 5. INSIGHTS BUSINESS CLÉS (5-8 points)
   - Chaque insight avec chiffres précis et interprétation
   - Impact business estimé (fort/moyen/faible)

## 6. RISQUES ET ANOMALIES
   - Valeurs aberrantes et leur signification
   - Données manquantes et biais potentiels
   - Points de vigilance pour la prise de décision

## 7. RECOMMANDATIONS ACTIONNABLES
   - Classées par priorité (P1/P2/P3)
   - Avec actions concrètes et responsables suggérés

## 8. PROCHAINES ÉTAPES
   - Analyses complémentaires suggérées
   - Données additionnelles à collecter
   - Modèles prédictifs pertinents à envisager

Sois TRÈS précis, cite les chiffres exacts, explique le POURQUOI de chaque observation.
L'objectif est d'aider des non-spécialistes à comprendre et agir sur leurs données.""")
    return "\n".join(lines)


def multi_model_analyze(eda, context, language, models, api_key):
    system = ("Tu es DataMind, analyste de données senior expert. Réponds en français avec précision."
              if language == "fr" else
              "You are DataMind, a senior data analyst. Reply in English with precision.")
    prompt = build_deep_analysis_prompt(eda, context, language)
    messages = [{"role":"user","content":prompt}]
    results = {}
    def _call(model):
        return model, call_openrouter(messages, model, api_key, system=system, max_tokens=3000)
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as ex:
        for m, text in ex.map(lambda m: _call(m), models):
            results[m] = text
    return results


def synthesize(analyses, language, api_key, model):
    valid = {k:v for k,v in analyses.items() if not v.startswith("❌")}
    if not valid:
        return "Aucune analyse valide."
    lang = "français" if language == "fr" else "English"
    text = "\n\n---\n\n".join(f"### Modèle {k} :\n{v}" for k,v in valid.items())
    prompt = (f"Tu as {len(valid)} analyses indépendantes du même jeu de données. "
              f"Crée un rapport de synthèse DÉFINITIF en {lang} qui :\n"
              "1. Consolide les consensus entre modèles\n"
              "2. Met en avant les insights divergents ou complémentaires\n"
              "3. Produit des recommandations finales hiérarchisées\n"
              "4. Inclut un guide d'interprétation pour non-spécialistes\n\n"
              f"ANALYSES À SYNTHÉTISER :\n{text}")
    result, used_model = call_with_fallback([{"role":"user","content":prompt}], model, api_key,
                            system="Expert en synthèse analytique. Sois exhaustif et pédagogique.",
                            max_tokens=4000)
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 3. INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="DataMind AI", page_icon="🧠", layout="wide")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

/* ══════════════════════════════════════════════
   DATAMIND CI — Palette Côte d'Ivoire
   🟠 Orange  #F77F00
   ⚪ Blanc   #FFFFFF
   🟢 Vert    #009A44
   Fond sombre : #0e1a0f (vert très sombre)
   ══════════════════════════════════════════════ */

* { font-family: 'Space Grotesk', sans-serif; }

/* Fond principal — vert nuit ivoirien */
.stApp {
    background: linear-gradient(160deg, #0a1a0b 0%, #0e1a0f 50%, #0f120a 100%);
    color: #f0ede6;
}

/* Sidebar — orange sombre */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a0e00 0%, #1f1200 100%) !important;
    border-right: 2px solid #F77F00 !important;
}
section[data-testid="stSidebar"] * { color: #f0ede6 !important; }
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 { color: #F77F00 !important; }

/* Onglets */
.stTabs [data-baseweb="tab-list"] {
    background: #1a1f0d;
    border-radius: 12px;
    padding: 5px;
    gap: 4px;
    border: 1px solid #2d4a1e;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: #8a9e7a !important;
    border-radius: 8px !important;
    font-weight: 500;
    transition: all 0.2s;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #F77F00, #e06b00) !important;
    color: white !important;
    box-shadow: 0 2px 8px rgba(247,127,0,0.4);
}

/* Cartes métriques */
.metric-box {
    background: linear-gradient(135deg, #1a2e10, #1f3512);
    border: 1px solid #2d5a1e;
    border-top: 3px solid #009A44;
    border-radius: 14px;
    padding: 1.3rem 1.5rem;
    text-align: center;
    transition: transform 0.2s, box-shadow 0.2s;
}
.metric-box:hover {
    transform: translateY(-3px);
    box-shadow: 0 6px 20px rgba(0,154,68,0.25);
}
.metric-val {
    font-size: 2.1rem;
    font-weight: 700;
    background: linear-gradient(135deg, #F77F00, #ffad4d);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    line-height: 1.1;
}
.metric-lbl {
    font-size: 0.72rem;
    color: #7aab7a;
    margin-top: 5px;
    text-transform: uppercase;
    letter-spacing: 0.07em;
}

/* Cartes insight — bordure verte CI */
.insight-card {
    background: linear-gradient(135deg, #162010, #1c2d12);
    border: 1px solid #2d5a1e;
    border-left: 4px solid #009A44;
    border-radius: 12px;
    padding: 1rem 1.3rem;
    margin-bottom: 0.9rem;
    box-shadow: 0 2px 8px rgba(0,154,68,0.1);
}

/* Cartes warning — orange CI */
.warn-card {
    background: linear-gradient(135deg, #1f1200, #261700);
    border-left: 4px solid #F77F00;
    border-radius: 10px;
    padding: 0.9rem 1.1rem;
    margin-bottom: 0.7rem;
    border: 1px solid #3d2200;
    border-left: 4px solid #F77F00;
}

/* Cartes succès — vert CI */
.ok-card {
    background: linear-gradient(135deg, #0d2210, #0f2912);
    border: 1px solid #1e4a15;
    border-left: 4px solid #009A44;
    border-radius: 10px;
    padding: 0.9rem 1.1rem;
    margin-bottom: 0.7rem;
}

/* Titres de section */
.section-title {
    font-size: 1.25rem;
    font-weight: 700;
    color: #f0ede6;
    margin: 1.5rem 0 0.8rem;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid transparent;
    border-image: linear-gradient(90deg, #F77F00, #009A44) 1;
    letter-spacing: 0.01em;
}

/* Tags modèles */
.model-tag {
    background: #1a2e10;
    border: 1px solid #2d5a1e;
    border-radius: 6px;
    padding: 3px 10px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.75rem;
    color: #7adb7a;
    display: inline-block;
    margin: 2px;
}
.free-badge {
    background: #F77F00;
    color: white;
    border-radius: 4px;
    padding: 1px 6px;
    font-size: 0.65rem;
    font-weight: 700;
    margin-left: 5px;
}

/* Boutons — orange CI */
.stButton > button {
    background: linear-gradient(135deg, #F77F00, #e06b00) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    letter-spacing: 0.02em !important;
    transition: all 0.2s !important;
    box-shadow: 0 3px 10px rgba(247,127,0,0.3) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #ff9020, #F77F00) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 16px rgba(247,127,0,0.4) !important;
}

/* Inputs */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div,
.stNumberInput > div > div > input {
    background: #1a2e10 !important;
    border: 1px solid #2d5a1e !important;
    border-radius: 8px !important;
    color: #f0ede6 !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #F77F00 !important;
    box-shadow: 0 0 0 2px rgba(247,127,0,0.2) !important;
}

/* Multiselect */
.stMultiSelect > div {
    background: #1a2e10 !important;
    border: 1px solid #2d5a1e !important;
    border-radius: 8px !important;
}

/* File uploader */
.stFileUploader > div {
    background: #1a2e10 !important;
    border: 2px dashed #009A44 !important;
    border-radius: 14px !important;
}

/* DataFrames */
div[data-testid="stDataFrame"] {
    border: 1px solid #2d5a1e;
    border-radius: 12px;
    overflow: hidden;
    box-shadow: 0 2px 12px rgba(0,154,68,0.1);
}

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0e1a0f; }
::-webkit-scrollbar-thumb { background: #F77F00; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #009A44; }

/* Toggle / Slider */
.stSlider > div > div > div { background: #F77F00 !important; }
.stCheckbox > label > div:first-child { border-color: #009A44 !important; }

/* Progress bar */
.stProgress > div > div { background: linear-gradient(90deg, #009A44, #F77F00) !important; }

/* Alerts */
.stAlert { border-radius: 10px !important; border: 1px solid #2d5a1e !important; }

/* Expander */
.streamlit-expanderHeader {
    background: #1a2e10 !important;
    border-radius: 8px !important;
    border: 1px solid #2d5a1e !important;
    color: #f0ede6 !important;
}

/* Flag banner */
.ci-flag {
    display: flex;
    height: 5px;
    border-radius: 3px;
    overflow: hidden;
    margin-bottom: 1rem;
}
.ci-orange { flex: 1; background: #F77F00; }
.ci-white  { flex: 1; background: #FFFFFF; }
.ci-green  { flex: 1; background: #009A44; }
</style>
""", unsafe_allow_html=True)

# Palette graphiques — Côte d'Ivoire
DARK = dict(
    plot_bgcolor="#0e1a0f",
    paper_bgcolor="#0e1a0f",
    font_color="#f0ede6",
    xaxis=dict(gridcolor="#1e3d15", color="#7aab7a", zerolinecolor="#2d5a1e"),
    yaxis=dict(gridcolor="#1e3d15", color="#7aab7a", zerolinecolor="#2d5a1e"),
)
# Palette couleurs CI pour les graphiques
CI_COLORS = ["#F77F00","#009A44","#ffad4d","#00c957","#ff6b00","#00a83a",
             "#ffc680","#66d98a","#cc6200","#007a30"]

# ── Header ───────────────────────────────────────────────────────────────────
# Bandeau drapeau CI
st.markdown('''<div class="ci-flag">
  <div class="ci-orange"></div>
  <div class="ci-white"></div>
  <div class="ci-green"></div>
</div>''', unsafe_allow_html=True)

st.markdown("""
<div style="padding:0.8rem 0 1rem; display:flex; align-items:center; gap:16px;">
  <div style="font-size:3rem; filter:drop-shadow(0 4px 8px rgba(247,127,0,0.4));">🧠</div>
  <div>
    <h1 style="margin:0; font-size:2.2rem; font-weight:700;
               background:linear-gradient(135deg,#F77F00,#ffad4d);
               -webkit-background-clip:text; -webkit-text-fill-color:transparent;
               background-clip:text;">
      DataMind AI Platform
    </h1>
    <p style="margin:2px 0 0; color:#7aab7a; font-size:0.9rem; letter-spacing:0.03em;">
      🇨🇮 Côte d'Ivoire · Big Data Analytics × IA Multi-Modèles × Rapport Détaillé
    </p>
  </div>
  <div style="margin-left:auto; text-align:right;">
    <span style="background:linear-gradient(135deg,#F77F00,#e06b00); color:white;
                 border-radius:20px; padding:4px 14px; font-size:0.8rem; font-weight:700;">
      100% Gratuit 🎉
    </span>
  </div>
</div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    api_key = st.text_input("🔑 Clé OpenRouter", value=os.getenv("OPENROUTER_API_KEY",""),
                             type="password", help="Compte gratuit sur openrouter.ai")
    if api_key:
        os.environ["OPENROUTER_API_KEY"] = api_key
        st.markdown('<div style="color:#3fb950;font-size:0.8rem;">✅ Clé configurée</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div style="color:#f85149;font-size:0.8rem;">⚠️ Clé manquante → openrouter.ai</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 🤖 Modèles Gratuits")

    # Description rapide des modèles
    MODEL_LABELS = {
        "openrouter/auto":                             "🔀 Auto (meilleur dispo)",
        "meta-llama/llama-3.3-70b-instruct:free":     "🦙 LLaMA 70B — Meilleur général",
        "meta-llama/llama-3.1-8b-instruct:free":      "🦙 LLaMA 8B — Rapide",
        "deepseek/deepseek-r1-0528:free":              "🧠 DeepSeek R1 — Raisonnement",
        "deepseek/deepseek-v3-0324:free":              "🧠 DeepSeek V3 — Analyse",
        "qwen/qwen3-14b:free":                         "🌏 Qwen3 14B — Français",
        "qwen/qwen3-8b:free":                          "🌏 Qwen3 8B — Léger",
        "qwen/qwen3-235b-a22b:free":                   "🌏 Qwen3 235B — Puissant",
        "google/gemma-3-27b-it:free":                  "🔵 Gemma 27B — Google",
        "google/gemma-3n-e4b-it:free":                 "🔵 Gemma E4B — Compact",
        "microsoft/phi-4-reasoning-plus:free":         "🪟 Phi-4 — Microsoft",
        "nvidia/llama-3.3-nemotron-super-49b-v1:free": "🟢 Nemotron 49B — NVIDIA",
        "openai/gpt-4o-mini-search-preview:free":      "⚪ GPT-4o mini — OpenAI",
        "mistralai/mistral-7b-instruct:free":          "💨 Mistral 7B — Français",
        "thudm/glm-4-32b:free":                        "🔷 GLM-4 32B — Multilingue",
        "moonshotai/moonlight-16b-a3b-instruct:free":  "🌙 Moonlight 16B",
    }
    model_display = [MODEL_LABELS.get(m, m.split("/")[-1]) for m in FREE_MODELS]
    model_map = dict(zip(model_display, FREE_MODELS))

    selected_display = st.multiselect(
        "Modèles (tous gratuits) :",
        options=model_display,
        default=model_display[:3],
        help="Conseil : sélectionnez 2-3 modèles. Le routeur Auto essaie le meilleur disponible."
    )
    selected_models = [model_map[d] for d in selected_display]

    for d in selected_display:
        st.markdown(f'<span class="model-tag">{d}<span class="free-badge">FREE</span></span>', unsafe_allow_html=True)

    # Bouton de diagnostic
    st.markdown("---")
    if st.button("🔍 Tester la connexion IA", use_container_width=True):
        api_test = os.getenv("OPENROUTER_API_KEY","")
        if not api_test:
            st.error("Clé API manquante")
        else:
            with st.spinner("Test..."):
                test_result = call_openrouter(
                    [{"role":"user","content":"Dis juste OK"}],
                    "meta-llama/llama-3.1-8b-instruct:free",
                    api_test, max_tokens=10, retries=1)
            if test_result.startswith("❌"):
                st.error(f"Connexion échouée : {test_result}")
                st.info("Vérifiez votre clé sur openrouter.ai → Keys")
            else:
                st.success("✅ Connexion IA opérationnelle ! 🇨🇮")

    st.markdown("---")
    st.markdown("### 🌍 Rapport")
    language = st.selectbox("Langue", ["fr","en"], format_func=lambda x:"🇫🇷 Français" if x=="fr" else "🇬🇧 English")
    context  = st.text_area("Contexte métier", placeholder="Ex: Données RH d'une entreprise ivoirienne...", height=80)

    st.markdown("---")
    st.markdown("### 🔍 Anomalies")
    detect_anom   = st.toggle("Activer détection", value=True)
    contamination = st.slider("Taux estimé", 0.01, 0.2, 0.05, 0.01) if detect_anom else 0.05

    st.markdown("---")
    st.caption("DataMind AI v3.0 | 100% Gratuit")

# ── Onglets ──────────────────────────────────────────────────────────────────
tabs = st.tabs(["📂 Données","📊 Statistiques","📈 Graphiques","🔬 Analyse Avancée","🔡 ACM","🤖 IA & Questions","🤖 ML Auto","📊 Excel Export","📁 Comparaison","📄 Rapport PDF","📧 Email","🗄️ Base de données"])

# ╔══════════════════╗
# ║  ONGLET 1 DONNÉES ║
# ╚══════════════════╝
with tabs[0]:
    st.markdown('<div class="section-title">Charger vos données</div>', unsafe_allow_html=True)
    c1, c2 = st.columns([3,1])
    with c1:
        uploaded = st.file_uploader("Glissez votre fichier", type=["csv","xlsx","xls","parquet","json"])
    with c2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📦 Exemple Iris", use_container_width=True):
            from sklearn.datasets import load_iris
            iris = load_iris()
            df_ex = pd.DataFrame(iris.data, columns=iris.feature_names)
            df_ex["species"] = pd.Categorical.from_codes(iris.target, iris.target_names)
            st.session_state.update({"df":df_ex,"eda":full_eda(df_ex),"source":"Iris"})
            st.success("✅ Iris chargé!")
        if st.button("📦 Exemple Ventes", use_container_width=True):
            rng = np.random.default_rng(42); n=600
            df_ex = pd.DataFrame({
                "date": pd.date_range("2023-01-01",periods=n,freq="D"),
                "ventes": rng.normal(50000,15000,n).clip(0),
                "quantite": rng.integers(5,300,n),
                "prix_unitaire": rng.uniform(20,800,n).round(2),
                "marge": rng.normal(0.25,0.08,n).clip(0.01,0.6),
                "region": rng.choice(["Abidjan","Bouaké","Daloa","Yamoussoukro","San-Pédro"],n),
                "categorie": rng.choice(["Électronique","Alimentaire","Textile","Cosmétique","Pharma"],n),
                "canal": rng.choice(["Boutique","En ligne","Grossiste","Direct"],n),
                "satisfaction": rng.integers(1,6,n),
            })
            st.session_state.update({"df":df_ex,"eda":full_eda(df_ex),"source":"Ventes exemple"})
            st.success("✅ Données de ventes chargées!")

    if uploaded:
        try:
            ext = Path(uploaded.name).suffix.lower()
            with st.spinner("Chargement et détection d'encodage..."):
                df_raw = None
                if ext == ".csv":
                    raw = uploaded.read()
                    for enc in ["utf-8","latin-1","windows-1252","iso-8859-1","utf-8-sig","cp1252"]:
                        try:
                            sample = raw[:2048].decode(enc, errors="ignore")
                            sep = ";" if sample.count(";") > sample.count(",") else ","
                            df_raw = pd.read_csv(io.BytesIO(raw), encoding=enc, sep=sep, on_bad_lines="skip")
                            st.caption(f"🔤 Encodage : **{enc}** | Séparateur : **`{sep}`**")
                            break
                        except: continue
                elif ext in (".xlsx",".xls"):
                    df_raw = pd.read_excel(uploaded)
                elif ext == ".parquet":
                    df_raw = pd.read_parquet(uploaded)
                elif ext == ".json":
                    df_raw = pd.read_json(uploaded)

            if df_raw is not None:
                df_c, rep = auto_clean(df_raw)
                st.session_state.update({"df":df_c,"df_raw":df_raw,"clean_report":rep,
                                          "eda":full_eda(df_c),"source":uploaded.name})
                st.success(f"✅ **{uploaded.name}** — {len(df_c):,} lignes × {df_c.shape[1]} colonnes")
        except Exception as e:
            st.error(f"Erreur : {e}")

    if "df" in st.session_state:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        ov  = eda["overview"]

        if detect_anom and "anomalie" not in df.columns:
            df = detect_anomalies(df, contamination)
            st.session_state["df"] = df

        st.markdown("<br>", unsafe_allow_html=True)
        cols5 = st.columns(5)
        for col, (val,lbl) in zip(cols5, [
            (f"{ov['rows']:,}","Lignes"), (str(ov["columns"]),"Colonnes"),
            (f"{ov['memory_mb']} Mo","Mémoire"), (f"{ov['missing_pct']}%","Manquants"),
            (str(int(df["anomalie"].sum())) if "anomalie" in df.columns else "—","Anomalies"),
        ]):
            col.markdown(f'<div class="metric-box"><div class="metric-val">{val}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

        if "clean_report" in st.session_state:
            with st.expander("🧹 Rapport de nettoyage"):
                r = st.session_state["clean_report"]
                st.write(f"**Doublons supprimés :** {r.get('doublons_supprimés',0)}")
                for col,act in r.get("valeurs_manquantes",{}).items():
                    st.write(f"• `{col}` : {act}")

        st.markdown('<div class="section-title">Aperçu</div>', unsafe_allow_html=True)
        n = st.slider("Lignes", 5, 200, 20)
        st.dataframe(df.head(n), use_container_width=True, height=380)

        with st.expander("📋 Types des colonnes"):
            td = pd.DataFrame({"Colonne":df.columns,"Type":df.dtypes.astype(str).values,
                                "Non-nulls":df.count().values,"Nulls":df.isna().sum().values,
                                "Uniques":[df[c].nunique() for c in df.columns]})
            st.dataframe(td, use_container_width=True)

# ╔═════════════════════════╗
# ║  ONGLET 2 STATISTIQUES  ║
# ╚═════════════════════════╝
with tabs[1]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        num_cols = eda["overview"]["numeric_cols"]
        cat_cols = eda["overview"]["categorical_cols"]

        st.markdown('<div class="section-title">Statistiques descriptives complètes</div>', unsafe_allow_html=True)
        if "descriptive_stats" in eda:
            st.dataframe(pd.DataFrame(eda["descriptive_stats"]).round(3), use_container_width=True)

        st.markdown('<div class="section-title">Analyse des distributions</div>', unsafe_allow_html=True)
        if dist := eda.get("distributions"):
            rows = []
            for col, d in dist.items():
                sk = d["skewness"]
                rows.append({
                    "Variable": col,
                    "Asymétrie": sk,
                    "Kurtosis": d["kurtosis"],
                    "Outliers (IQR)": d["outliers_iqr"],
                    "Distribution": "✅ Normale" if d["is_normal"] else "⚠️ Non-normale",
                    "Interprétation": ("⬅️ Asymétrie gauche" if sk < -1 else
                                       "➡️ Asymétrie droite" if sk > 1 else "↔️ Symétrique"),
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True)

        st.markdown('<div class="section-title">Corrélations</div>', unsafe_allow_html=True)
        strong = eda.get("correlations",{}).get("strong_pairs",[])
        if strong:
            rows2 = []
            for p in strong:
                r = p["r"]
                rows2.append({
                    "Variable 1": p["col1"], "Variable 2": p["col2"],
                    "Corrélation r": r,
                    "Force": "🔴 Très forte" if abs(r)>0.8 else "🟠 Forte" if abs(r)>0.65 else "🟡 Modérée",
                    "Direction": "📈 Positive" if r>0 else "📉 Négative",
                    "Interprétation": f"Quand {p['col1']} augmente, {p['col2']} {'augmente' if r>0 else 'diminue'}",
                })
            st.dataframe(pd.DataFrame(rows2), use_container_width=True)
        else:
            st.info("Aucune corrélation forte détectée (|r| > 0.4).")

        if cat_cols:
            st.markdown('<div class="section-title">Variables catégorielles</div>', unsafe_allow_html=True)
            for col, info in eda.get("categorical_analysis",{}).items():
                with st.expander(f"📌 **{col}** — {info['unique_count']} valeurs uniques"):
                    df_cat = pd.DataFrame({
                        "Valeur": list(info["top_values"].keys()),
                        "Fréquence": list(info["top_values"].values()),
                        "Pourcentage": [f"{v}%" for v in info["top_pct"].values()],
                    })
                    st.dataframe(df_cat, use_container_width=True)

# ╔══════════════════════╗
# ║  ONGLET 3 GRAPHIQUES ║
# ╚══════════════════════╝
with tabs[2]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        num_cols = eda["overview"]["numeric_cols"]
        cat_cols = eda["overview"]["categorical_cols"]
        dt_cols  = eda["overview"]["datetime_cols"]

        # ─ Distributions ─────────────────────────────────────────────────────
        if num_cols:
            st.markdown('<div class="section-title">📊 Distributions</div>', unsafe_allow_html=True)
            g1,g2 = st.columns(2)
            col_d = g1.selectbox("Variable :", num_cols, key="d1")
            with g1:
                fig = px.histogram(df, x=col_d, nbins=50, marginal="box",
                                   color_discrete_sequence=["#F77F00"],
                                   title=f"Histogramme + Boxplot — {col_d}")
                fig.update_layout(**DARK, title_font_size=13)
                st.plotly_chart(fig, use_container_width=True)
            with g2:
                fig = go.Figure()
                fig.add_trace(go.Violin(y=df[col_d].dropna(), box_visible=True, meanline_visible=True,
                                        fillcolor="#F77F00", line_color="#ffad4d", opacity=0.7, name=col_d))
                fig.update_layout(title=f"Violin — {col_d}", **DARK, title_font_size=13, showlegend=False)
                st.plotly_chart(fig, use_container_width=True)

        # ─ Toutes distributions côte à côte ──────────────────────────────────
        if num_cols:
            st.markdown('<div class="section-title">📦 Boxplots comparatifs</div>', unsafe_allow_html=True)
            df_norm = df[num_cols].copy()
            for c in df_norm.columns:
                r = df_norm[c].max() - df_norm[c].min()
                if r > 0: df_norm[c] = (df_norm[c] - df_norm[c].min()) / r
            fig = go.Figure()
            for col in num_cols[:12]:
                fig.add_trace(go.Box(y=df_norm[col].dropna(), name=col, boxpoints="outliers",
                                     marker_color="#F77F00", line_color="#ffad4d"))
            fig.update_layout(title="Boxplots normalisés (détection outliers)", **DARK,
                              title_font_size=13, showlegend=False)
            st.plotly_chart(fig, use_container_width=True)

        # ─ Heatmap corrélation ────────────────────────────────────────────────
        if len(num_cols) >= 2:
            st.markdown('<div class="section-title">🌡️ Heatmap de corrélation</div>', unsafe_allow_html=True)
            corr_m = df[num_cols].corr()
            fig = px.imshow(corr_m, text_auto=".2f", color_continuous_scale=["#F77F00","#0e1a0f","#009A44"],
                            zmin=-1, zmax=1, aspect="auto", title="Matrice de corrélation")
            fig.update_layout(**DARK, title_font_size=13)
            st.plotly_chart(fig, use_container_width=True)

        # ─ Scatter matrix ────────────────────────────────────────────────────
        if len(num_cols) >= 3:
            st.markdown('<div class="section-title">🔵 Scatter Matrix (Pairplot)</div>', unsafe_allow_html=True)
            cols_sel = st.multiselect("Variables :", num_cols, default=num_cols[:min(4,len(num_cols))], key="pm")
            if len(cols_sel) >= 2:
                color_pm = st.selectbox("Couleur :", ["—"]+cat_cols, key="pmc")
                fig = px.scatter_matrix(df, dimensions=cols_sel,
                                         color=None if color_pm=="—" else color_pm,
                                         color_discrete_sequence=CI_COLORS,
                                         title="Scatter Matrix")
                fig.update_traces(diagonal_visible=False, showupperhalf=False, marker=dict(size=3, opacity=0.6))
                fig.update_layout(**DARK, title_font_size=13, height=600)
                st.plotly_chart(fig, use_container_width=True)

        # ─ Scatter simple ────────────────────────────────────────────────────
        if len(num_cols) >= 2:
            st.markdown('<div class="section-title">🎯 Nuage de points</div>', unsafe_allow_html=True)
            sc1,sc2,sc3 = st.columns(3)
            xc = sc1.selectbox("Axe X",num_cols,index=0,key="sx")
            yc = sc2.selectbox("Axe Y",num_cols,index=min(1,len(num_cols)-1),key="sy")
            cc = sc3.selectbox("Couleur",["—"]+cat_cols,key="sc")
            fig = px.scatter(df, x=xc, y=yc, color=None if cc=="—" else cc,
                             color_discrete_sequence=CI_COLORS,
                             opacity=0.65, title=f"{xc} vs {yc}")
            fig.update_layout(**DARK, title_font_size=13)
            st.plotly_chart(fig, use_container_width=True)

        # ─ Catégorielles ─────────────────────────────────────────────────────
        if cat_cols:
            st.markdown('<div class="section-title">📌 Catégorielles</div>', unsafe_allow_html=True)
            cs = st.selectbox("Variable :", cat_cols, key="cs")
            vc = df[cs].value_counts().head(15)
            ca1,ca2 = st.columns(2)
            with ca1:
                fig = px.bar(x=vc.values, y=vc.index, orientation="h",
                             color=vc.values, color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                             title=f"Fréquences — {cs}")
                fig.update_layout(**DARK, title_font_size=13, showlegend=False, coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)
            with ca2:
                fig = px.pie(values=vc.values, names=vc.index, title=f"Répartition — {cs}",
                             color_discrete_sequence=CI_COLORS, hole=0.35)
                fig.update_layout(paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                                  font_color="#e6edf3", title_font_size=13)
                st.plotly_chart(fig, use_container_width=True)

            # Croisement catégorielle × numérique
            if num_cols:
                st.markdown('<div class="section-title">📊 Croisement catégorielle × numérique</div>', unsafe_allow_html=True)
                cv1,cv2 = st.columns(2)
                cat_x = cv1.selectbox("Catégorie :", cat_cols, key="cx")
                num_y = cv2.selectbox("Valeur numérique :", num_cols, key="cy")
                fig = px.box(df, x=cat_x, y=num_y, color=cat_x,
                             color_discrete_sequence=CI_COLORS,
                             title=f"Distribution de {num_y} par {cat_x}")
                fig.update_layout(**DARK, title_font_size=13, showlegend=False)
                st.plotly_chart(fig, use_container_width=True)

                fig2 = px.violin(df, x=cat_x, y=num_y, color=cat_x, box=True,
                                 color_discrete_sequence=CI_COLORS,
                                 title=f"Violin {num_y} par {cat_x}")
                fig2.update_layout(**DARK, title_font_size=13, showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

        # ─ Séries temporelles ────────────────────────────────────────────────
        if dt_cols and num_cols:
            st.markdown('<div class="section-title">📅 Séries temporelles</div>', unsafe_allow_html=True)
            dt1,dt2,dt3 = st.columns(3)
            dtc = dt1.selectbox("Date :", dt_cols, key="dtc")
            vc2 = dt2.selectbox("Valeur :", num_cols, key="vc2")
            grp = dt3.selectbox("Grouper par :", ["—"]+cat_cols, key="grp")
            ts  = df[[dtc,vc2]+([grp] if grp!="—" else [])].dropna().sort_values(dtc)
            fig = px.line(ts, x=dtc, y=vc2, color=None if grp=="—" else grp,
                          title=f"Évolution de {vc2}", color_discrete_sequence=CI_COLORS)
            fig.update_layout(**DARK, title_font_size=13)
            fig.update_traces(line=dict(width=2))
            st.plotly_chart(fig, use_container_width=True)

        # ─ Anomalies ─────────────────────────────────────────────────────────
        if "anomalie" in df.columns and len(num_cols) >= 2:
            st.markdown('<div class="section-title">🚨 Anomalies détectées</div>', unsafe_allow_html=True)
            an1,an2 = st.columns(2)
            ax = an1.selectbox("Axe X :", num_cols, index=0, key="ax")
            ay = an2.selectbox("Axe Y :", num_cols, index=min(1,len(num_cols)-1), key="ay")
            fig = px.scatter(df, x=ax, y=ay, color="anomalie",
                             color_discrete_map={True:"#f85149",False:"#1f6feb"},
                             title="Anomalies (rouge) vs Normaux (bleu)",
                             opacity=0.7, size_max=8)
            fig.update_layout(**DARK, title_font_size=13)
            st.plotly_chart(fig, use_container_width=True)
            n_a = int(df["anomalie"].sum())
            st.markdown(f'<div class="warn-card">🔴 <strong>{n_a} anomalies</strong> sur {len(df):,} lignes ({n_a/len(df)*100:.1f}%) — Examinez ces lignes attentivement</div>', unsafe_allow_html=True)
            with st.expander("Voir les lignes anomales"):
                st.dataframe(df[df["anomalie"]==True].drop(columns=["anomalie","score_anomalie"],errors="ignore").head(50), use_container_width=True)

# ╔══════════════════════════════╗
# ║  ONGLET 4 ANALYSE AVANCÉE   ║
# ╚══════════════════════════════╝
with tabs[3]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        num_cols = eda["overview"]["numeric_cols"]
        cat_cols = eda["overview"]["categorical_cols"]

        # ─ ACP / PCA ─────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">🔬 Analyse en Composantes Principales (ACP)</div>', unsafe_allow_html=True)
        if len(num_cols) >= 2:
            df_pca, explained = compute_pca(df)
            if df_pca is not None:
                p1,p2 = st.columns(2)
                with p1:
                    color_pca = st.selectbox("Couleur :", ["—"]+cat_cols, key="pcac")
                    if "anomalie" in df.columns:
                        color_pca = st.selectbox("Ou colorier par anomalie ?", ["Non","Oui"], key="pcaa")

                fig_pca = px.scatter(
                    df_pca, x="PC1", y="PC2" if "PC2" in df_pca.columns else "PC1",
                    title=f"ACP — PC1({explained[0]*100:.1f}%) vs PC2({explained[1]*100:.1f}% variance expliquée)" if len(explained)>1 else "ACP",
                    color_discrete_sequence=["#F77F00"], opacity=0.7
                )
                fig_pca.update_layout(**DARK, title_font_size=12)
                st.plotly_chart(fig_pca, use_container_width=True)

                # Variance expliquée
                fig_var = px.bar(x=[f"PC{i+1}" for i in range(len(explained))],
                                  y=[e*100 for e in explained],
                                  title="Variance expliquée par composante (%)",
                                  color_discrete_sequence=["#F77F00"])
                fig_var.update_layout(**DARK, title_font_size=12, xaxis_title="Composante", yaxis_title="Variance (%)")
                st.plotly_chart(fig_var, use_container_width=True)
        else:
            st.info("L'ACP nécessite au moins 2 variables numériques.")

        # ─ Analyse de concentration ───────────────────────────────────────────
        if num_cols:
            st.markdown('<div class="section-title">📐 Courbe de Lorenz (Concentration)</div>', unsafe_allow_html=True)
            lc = st.selectbox("Variable :", num_cols, key="lorenz")
            vals = df[lc].dropna().sort_values().values
            vals = vals[vals > 0]
            if len(vals) > 0:
                cum = np.cumsum(vals) / vals.sum()
                pop = np.arange(1, len(vals)+1) / len(vals)
                gini = 1 - 2*np.trapezoid(cum, pop) if hasattr(np, "trapezoid") else np.trapz(cum, pop)
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=[0]+list(pop), y=[0]+list(cum), fill="tozeroy",
                                          name="Lorenz", line=dict(color="#F77F00",width=2)))
                fig.add_trace(go.Scatter(x=[0,1], y=[0,1], name="Égalité parfaite",
                                          line=dict(color="#009A44", dash="dash", width=1)))
                fig.update_layout(title=f"Courbe de Lorenz — {lc} | Gini={gini:.3f}", **DARK,
                                   title_font_size=12, xaxis_title="Proportion cumulée", yaxis_title="Part cumulée")
                st.plotly_chart(fig, use_container_width=True)
                gini_label = "Très inégale" if gini>0.6 else "Inégale" if gini>0.4 else "Modérée" if gini>0.2 else "Équilibrée"
                st.markdown(f'<div class="insight-card">📊 Indice de Gini : <strong>{gini:.3f}</strong> — Distribution <strong>{gini_label}</strong></div>', unsafe_allow_html=True)

        # ─ Tests statistiques ────────────────────────────────────────────────
        if len(num_cols) >= 2:
            st.markdown('<div class="section-title">🧪 Tests Statistiques</div>', unsafe_allow_html=True)
            t1,t2 = st.columns(2)
            col_a = t1.selectbox("Variable A :", num_cols, index=0, key="ta")
            col_b = t2.selectbox("Variable B :", num_cols, index=min(1,len(num_cols)-1), key="tb")

            a = df[col_a].dropna()
            b = df[col_b].dropna()

            try:
                t_stat, t_pval = stats.ttest_ind(a,b)
                _, norm_a = stats.shapiro(a.sample(min(5000,len(a)), random_state=42))
                _, norm_b = stats.shapiro(b.sample(min(5000,len(b)), random_state=42))
                corr_val, corr_p = stats.pearsonr(df[[col_a,col_b]].dropna()[col_a], df[[col_a,col_b]].dropna()[col_b])

                res = pd.DataFrame([
                    ["Test t de Student (moyennes égales ?)", f"t={t_stat:.3f}", f"p={t_pval:.4f}", "✅ Moyennes similaires" if t_pval>0.05 else "⚠️ Moyennes différentes"],
                    [f"Normalité {col_a} (Shapiro-Wilk)", "—", f"p={norm_a:.4f}", "✅ Normale" if norm_a>0.05 else "⚠️ Non-normale"],
                    [f"Normalité {col_b} (Shapiro-Wilk)", "—", f"p={norm_b:.4f}", "✅ Normale" if norm_b>0.05 else "⚠️ Non-normale"],
                    ["Corrélation de Pearson", f"r={corr_val:.3f}", f"p={corr_p:.4f}", "✅ Non significative" if corr_p>0.05 else "⚠️ Significative"],
                ], columns=["Test","Statistique","p-valeur","Interprétation"])
                st.dataframe(res, use_container_width=True)
                st.caption("Règle : p-valeur < 0.05 → résultat statistiquement significatif")
            except Exception as e:
                st.warning(f"Calcul impossible : {e}")

        # ─ Distribution groupée ───────────────────────────────────────────────
        if cat_cols and num_cols:
            st.markdown('<div class="section-title">📊 Analyse Groupée (Group By)</div>', unsafe_allow_html=True)
            gb1,gb2,gb3 = st.columns(3)
            grp_col = gb1.selectbox("Grouper par :", cat_cols, key="gb1")
            agg_col = gb2.selectbox("Variable :", num_cols, key="gb2")
            agg_fn  = gb3.selectbox("Agrégation :", ["mean","median","sum","count","std","min","max"], key="gb3")
            grouped = df.groupby(grp_col)[agg_col].agg(agg_fn).reset_index().sort_values(agg_col, ascending=False)
            fig = px.bar(grouped, x=grp_col, y=agg_col, color=agg_col,
                          color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                          title=f"{agg_fn.upper()} de {agg_col} par {grp_col}")
            fig.update_layout(**DARK, title_font_size=12, showlegend=False, coloraxis_showscale=False)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(grouped.rename(columns={agg_col:f"{agg_fn}({agg_col})"}), use_container_width=True)

# ╔══════════════════════════════════════╗
# ║  ONGLET 5 — ANALYSE ACM              ║
# ╚══════════════════════════════════════╝
with tabs[4]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        cat_cols = eda["overview"]["categorical_cols"]

        st.markdown('''<div class="insight-card">
<strong>🔡 Analyse des Correspondances Multiples (ACM)</strong><br>
L'ACM est l'équivalent de l'ACP pour les <strong>variables catégorielles</strong>.
Elle réduit la dimensionnalité et révèle les associations cachées entre modalités.
</div>''', unsafe_allow_html=True)

        if len(cat_cols) < 2:
            st.warning("⚠️ L'ACM nécessite au moins 2 variables catégorielles.")
        else:
            st.markdown('<div class="section-title">1. Sélection des variables</div>', unsafe_allow_html=True)
            acm_cols = st.multiselect("Variables catégorielles :", cat_cols,
                                       default=cat_cols[:min(5,len(cat_cols))], key="acm_vars")
            max_mod = st.slider("Modalités max par variable", 3, 20, 8, key="acm_mod")

            if len(acm_cols) >= 2:
                # ── Encodage disjonctif ───────────────────────────────────────
                dummies_list = []
                for col in acm_cols:
                    top_vals = df[col].value_counts().head(max_mod).index
                    df_col = df[col].where(df[col].isin(top_vals), other="Autre")
                    dummies_list.append(pd.get_dummies(df_col, prefix=col))
                X_dummies = pd.concat(dummies_list, axis=1).fillna(0).astype(float)
                n_rows, n_cols_d = X_dummies.shape

                st.markdown('<div class="section-title">2. Résumé du tableau disjonctif</div>', unsafe_allow_html=True)
                m1,m2,m3 = st.columns(3)
                m1.markdown(f'<div class="metric-box"><div class="metric-val">{n_rows:,}</div><div class="metric-lbl">Individus</div></div>', unsafe_allow_html=True)
                m2.markdown(f'<div class="metric-box"><div class="metric-val">{len(acm_cols)}</div><div class="metric-lbl">Variables actives</div></div>', unsafe_allow_html=True)
                m3.markdown(f'<div class="metric-box"><div class="metric-val">{n_cols_d}</div><div class="metric-lbl">Modalités totales</div></div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)

                with st.expander("📋 Tableau disjonctif (20 premières lignes)"):
                    st.dataframe(X_dummies.head(20), use_container_width=True)

                try:
                    # ── Calcul ACM par SVD ────────────────────────────────────
                    Z = X_dummies.values
                    N = Z.sum()
                    r = Z.sum(axis=1) / N
                    c = Z.sum(axis=0) / N
                    P = Z / N
                    D_r_inv = np.diag(1.0 / (r * N + 1e-10))
                    D_c_inv = np.diag(1.0 / (c + 1e-10))
                    S = D_r_inv @ (P - np.outer(r, c)) @ D_c_inv
                    U, sigma, Vt = np.linalg.svd(S, full_matrices=False)
                    n_comp = min(5, len(sigma))
                    sigma   = sigma[:n_comp]
                    U       = U[:, :n_comp]
                    Vt      = Vt[:n_comp, :]
                    coord_ind = (D_r_inv @ P @ D_c_inv @ Vt.T) * sigma
                    coord_mod = (D_c_inv @ Vt.T) * sigma
                    eigenvalues   = sigma ** 2
                    total_inertia = eigenvalues.sum()
                    inertia_pct   = eigenvalues / total_inertia * 100 if total_inertia > 0 else eigenvalues * 0
                    axes_labels   = [f"Axe {i+1}" for i in range(n_comp)]
                    mod_names     = X_dummies.columns.tolist()

                    # ── Éboulis des valeurs propres ───────────────────────────
                    st.markdown('<div class="section-title">3. Éboulis des valeurs propres</div>', unsafe_allow_html=True)
                    fig_sc = go.Figure()
                    fig_sc.add_trace(go.Bar(x=axes_labels, y=inertia_pct.tolist(),
                                             marker_color="#F77F00", name="Inertie (%)"))
                    fig_sc.add_trace(go.Scatter(x=axes_labels, y=np.cumsum(inertia_pct).tolist(),
                                                 mode="lines+markers", name="Cumulée (%)",
                                                 line=dict(color="#009A44",width=2), marker=dict(size=8, color="#F77F00")))
                    fig_sc.update_layout(title="Éboulis — Inertie par axe factoriel", **DARK,
                                          xaxis_title="Axe", yaxis_title="Inertie (%)",
                                          legend=dict(bgcolor="#161b22"))
                    st.plotly_chart(fig_sc, use_container_width=True)

                    eigen_df = pd.DataFrame({
                        "Axe": axes_labels,
                        "Valeur propre": eigenvalues.round(5).tolist(),
                        "Inertie (%)": inertia_pct.round(2).tolist(),
                        "Cumulée (%)": np.cumsum(inertia_pct).round(2).tolist(),
                        "Qualité": ["⭐ Très important" if p>20 else "✅ Important" if p>10 else "ℹ️ Secondaire" for p in inertia_pct],
                    })
                    st.dataframe(eigen_df, use_container_width=True)

                    # ── Sélection axes ────────────────────────────────────────
                    st.markdown('<div class="section-title">4. Plan factoriel</div>', unsafe_allow_html=True)
                    c1,c2,c3 = st.columns(3)
                    ax1_i = c1.selectbox("Axe horizontal :", axes_labels, index=0, key="acm_h")
                    ax2_i = c2.selectbox("Axe vertical :", axes_labels, index=min(1,n_comp-1), key="acm_v")
                    color_ind = c3.selectbox("Couleur individus :", ["—"]+cat_cols, key="acm_ci")
                    ax1 = int(ax1_i.split()[-1])-1
                    ax2 = int(ax2_i.split()[-1])-1

                    # ── Plan individus ────────────────────────────────────────
                    df_ind = pd.DataFrame({ax1_i: coord_ind[:len(df),ax1], ax2_i: coord_ind[:len(df),ax2]})
                    if color_ind != "—" and color_ind in df.columns:
                        df_ind["Groupe"] = df[color_ind].values[:len(df_ind)]
                        fig_ind = px.scatter(df_ind, x=ax1_i, y=ax2_i, color="Groupe",
                                              color_discrete_sequence=CI_COLORS,
                                              title=f"Individus — {ax1_i} vs {ax2_i}", opacity=0.6)
                    else:
                        fig_ind = px.scatter(df_ind, x=ax1_i, y=ax2_i,
                                              color_discrete_sequence=["#F77F00"],
                                              title=f"Individus — {ax1_i} vs {ax2_i}", opacity=0.5)
                    fig_ind.add_hline(y=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_ind.add_vline(x=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_ind.update_layout(**DARK, title_font_size=13,
                                           xaxis_title=f"{ax1_i} ({inertia_pct[ax1]:.1f}%)",
                                           yaxis_title=f"{ax2_i} ({inertia_pct[ax2]:.1f}%)")
                    fig_ind.update_traces(marker=dict(size=5))

                    # ── Plan modalités ────────────────────────────────────────
                    df_mod = pd.DataFrame({
                        ax1_i: coord_mod[:len(mod_names),ax1],
                        ax2_i: coord_mod[:len(mod_names),ax2],
                        "Modalité": mod_names,
                        "Variable": [m.rsplit("_",1)[0] for m in mod_names],
                    })
                    fig_mod = px.scatter(df_mod, x=ax1_i, y=ax2_i, color="Variable", text="Modalité",
                                          color_discrete_sequence=CI_COLORS,
                                          title=f"Modalités — {ax1_i} vs {ax2_i}")
                    fig_mod.update_traces(marker=dict(size=11,opacity=0.9),
                                           textposition="top center", textfont=dict(size=8,color="#e6edf3"))
                    fig_mod.add_hline(y=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_mod.add_vline(x=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_mod.update_layout(**DARK, title_font_size=13, height=520,
                                           xaxis_title=f"{ax1_i} ({inertia_pct[ax1]:.1f}%)",
                                           yaxis_title=f"{ax2_i} ({inertia_pct[ax2]:.1f}%)")

                    # Affichage côte à côte
                    pi1, pi2 = st.columns(2)
                    with pi1:
                        st.plotly_chart(fig_ind, use_container_width=True)
                    with pi2:
                        st.plotly_chart(fig_mod, use_container_width=True)

                    # ── Biplot ────────────────────────────────────────────────
                    st.markdown('<div class="section-title">5. Biplot ACM (individus + modalités)</div>', unsafe_allow_html=True)
                    fig_bi = go.Figure()
                    fig_bi.add_trace(go.Scatter(
                        x=coord_ind[:len(df),ax1], y=coord_ind[:len(df),ax2],
                        mode="markers", name="Individus",
                        marker=dict(size=4, color="#1f6feb", opacity=0.35)))
                    pal = px.colors.qualitative.Plotly
                    for i_v, var in enumerate(df_mod["Variable"].unique()):
                        sub = df_mod[df_mod["Variable"]==var]
                        label = sub["Modalité"].str.replace(f"{var}_","",regex=False)
                        fig_bi.add_trace(go.Scatter(
                            x=sub[ax1_i], y=sub[ax2_i], mode="markers+text",
                            name=var, text=label, textposition="top center",
                            textfont=dict(size=8, color="#e6edf3"),
                            marker=dict(size=13, color=pal[i_v%len(pal)], symbol="diamond",
                                        opacity=0.95, line=dict(width=1,color="white"))))
                    fig_bi.add_hline(y=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_bi.add_vline(x=0,line_dash="dash",line_color="#30363d",line_width=1)
                    fig_bi.update_layout(
                        title=f"Biplot ACM — {ax1_i} ({inertia_pct[ax1]:.1f}%) × {ax2_i} ({inertia_pct[ax2]:.1f}%)",
                        **DARK, title_font_size=13, height=620,
                        xaxis_title=f"{ax1_i} ({inertia_pct[ax1]:.1f}%)",
                        yaxis_title=f"{ax2_i} ({inertia_pct[ax2]:.1f}%)",
                        legend=dict(bgcolor="#161b22",bordercolor="#30363d",font=dict(size=9)))
                    st.plotly_chart(fig_bi, use_container_width=True)

                    # ── Contributions ─────────────────────────────────────────
                    st.markdown('<div class="section-title">6. Contributions des modalités</div>', unsafe_allow_html=True)
                    contrib_ax = st.selectbox("Axe :", axes_labels, key="acm_ctb")
                    ctb_idx = int(contrib_ax.split()[-1])-1
                    contrib = coord_mod[:len(mod_names),ctb_idx]**2
                    contrib_pct = contrib/contrib.sum()*100
                    ctb_df = pd.DataFrame({
                        "Modalité": mod_names, "Variable":[m.rsplit("_",1)[0] for m in mod_names],
                        "Contribution (%)": contrib_pct.round(2),
                    }).sort_values("Contribution (%)", ascending=False).head(20)
                    fig_ctb = px.bar(ctb_df, x="Contribution (%)", y="Modalité", color="Variable",
                                      orientation="h", color_discrete_sequence=CI_COLORS,
                                      title=f"Contributions — {contrib_ax}")
                    fig_ctb.add_vline(x=100/len(mod_names), line_dash="dash", line_color="#F77F00",
                                       line_width=1.5, annotation_text="Seuil moyen")
                    fig_ctb.update_layout(**DARK, title_font_size=12, height=500)
                    st.plotly_chart(fig_ctb, use_container_width=True)

                    # ── Tableau de contingence + Khi-deux ─────────────────────
                    st.markdown('<div class="section-title">7. Tableau de contingence & Test Khi-deux</div>', unsafe_allow_html=True)
                    k1,k2 = st.columns(2)
                    var_r = k1.selectbox("Lignes :", acm_cols, index=0, key="khi_r")
                    var_c = k2.selectbox("Colonnes :", acm_cols, index=min(1,len(acm_cols)-1), key="khi_c")
                    if var_r != var_c:
                        from scipy.stats import chi2_contingency
                        ct = pd.crosstab(df[var_r], df[var_c])
                        chi2, p_chi, dof, _ = chi2_contingency(ct)
                        cv = np.sqrt(chi2/(ct.values.sum()*(min(ct.shape)-1)))
                        sig = p_chi < 0.05
                        force = "forte" if cv>0.3 else "modérée" if cv>0.15 else "faible"
                        cls = "ok-card" if sig else "insight-card"
                        assoc_txt = "✅ Association significative" if sig else "⚪ Pas d'association"
                        st.markdown(f'<div class="{cls}">🧪 χ²={chi2:.2f} | p={p_chi:.4f} | V de Cramér={cv:.3f} | {assoc_txt} ({force})</div>', unsafe_allow_html=True)
                        fig_ct = px.imshow(ct, text_auto=True, color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                                            title=f"{var_r} × {var_c}")
                        fig_ct.update_layout(**DARK, title_font_size=12)
                        st.plotly_chart(fig_ct, use_container_width=True)
                        st.dataframe(ct, use_container_width=True)
                    else:
                        st.info("Choisissez deux variables différentes.")

                    # Guide d'interpretation
                    guide_html = (
                        '<div class="section-title">📖 Guide d\'interprétation ACM</div>'
                    )
                    st.markdown(guide_html, unsafe_allow_html=True)
                    guide_text = (
                        "<div class=\"insight-card\">"
                        "<strong>Comment lire les graphiques ACM :</strong><br><br>"
                        "🔵 <strong>Individus proches</strong> → Profils similaires<br>"
                        "🔴 <strong>Modalités proches</strong> → Catégories qui coexistent souvent<br>"
                        "📐 <strong>Modalité proche d\'un groupe</strong> → Ces individus ont cette modalité<br>"
                        "📊 <strong>Axe 1</strong> → Principale source de variabilité<br>"
                        "📊 <strong>Axe 2</strong> → Deuxième source, indépendante de l\'axe 1<br>"
                        "⚡ <strong>Contribution &gt; seuil jaune</strong> → Modalité structurante<br>"
                        "🧪 <strong>V de Cramér &gt; 0.3</strong> → Forte association entre variables<br>"
                        "💡 <strong>Modalités opposées</strong> → Groupes qui s\'excluent mutuellement"
                        "</div>"
                    )
                    st.markdown(guide_text, unsafe_allow_html=True)

                except Exception as e:
                    import traceback
                    st.error(f"Erreur calcul ACM : {e}")
                    st.code(traceback.format_exc())
            else:
                st.info("Sélectionnez au moins 2 variables catégorielles.")


# ╔══════════════════════════════╗
# ║  ONGLET 5 — IA & QUESTIONS  ║
# ╚══════════════════════════════╝
with tabs[5]:
    if "eda" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    elif not os.getenv("OPENROUTER_API_KEY",""):
        st.warning("🔑 Ajoutez votre clé OpenRouter dans la barre latérale.")
    elif not selected_models:
        st.warning("Sélectionnez au moins un modèle dans la barre latérale.")
    else:
        eda = st.session_state["eda"]
        api = os.getenv("OPENROUTER_API_KEY","")
        model1 = selected_models[0]

        # ─ Bloc 1 : L'IA pose ses propres questions ──────────────────────────
        st.markdown('<div class="section-title">🤔 L\'IA interroge vos données</div>', unsafe_allow_html=True)
        st.markdown('<div class="insight-card">L\'IA va <strong>générer elle-même des questions analytiques</strong> pertinentes sur vos données, puis y <strong>répondre avec les statistiques réelles</strong>.</div>', unsafe_allow_html=True)

        if st.button("🧠 Générer les questions et réponses automatiques", type="primary", use_container_width=True):
            # Étape 1 : génération des questions
            with st.spinner(f"🤔 {model1.split('/')[-1].replace(':free','')} formule ses questions..."):
                questions = generate_ai_questions(eda, context, language, model1, api)
                st.session_state["ai_questions"] = questions

            if questions.startswith("❌"):
                st.error(f"Impossible de générer les questions : {questions}")
                st.info("💡 Conseil : Vérifiez votre clé OpenRouter et essayez un autre modèle dans la barre latérale.")
            else:
                st.success("✅ Questions générées !")
                # Étape 2 : réponses aux questions
                with st.spinner("💡 L'IA analyse vos données et répond..."):
                    answers = answer_ai_questions(eda, questions, context, language, model1, api)
                    st.session_state["ai_answers"] = answers

                if answers.startswith("❌"):
                    st.error(f"Impossible de répondre aux questions : {answers}")
                else:
                    st.success("✅ Auto-analyse terminée !")

        if "ai_questions" in st.session_state:
            q1,q2 = st.columns(2)
            with q1:
                st.markdown("**❓ Questions générées par l'IA :**")
                st.markdown(f'<div class="insight-card">{st.session_state["ai_questions"]}</div>', unsafe_allow_html=True)
            with q2:
                st.markdown("**💡 Réponses basées sur vos données :**")
                if "ai_answers" in st.session_state:
                    st.markdown(f'<div class="ok-card">{st.session_state["ai_answers"]}</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ─ Bloc 2 : Analyse multi-modèles profonde ───────────────────────────
        st.markdown('<div class="section-title">🚀 Analyse Multi-Modèles Approfondie</div>', unsafe_allow_html=True)
        if st.button("🔬 Lancer l'analyse complète", type="primary", use_container_width=True):
            prog = st.progress(0, "Initialisation...")
            with st.spinner(f"Analyse avec {len(selected_models)} modèle(s)..."):
                prog.progress(20, "Envoi aux modèles IA...")
                analyses = multi_model_analyze(eda, context, language, selected_models, api)
                prog.progress(70, "Synthèse en cours...")
                valid_analyses = {k:v for k,v in analyses.items() if not v.startswith("❌")}
                if valid_analyses:
                    synthesis = synthesize(valid_analyses, language, api, model1)
                else:
                    synthesis = "❌ Aucune analyse valide — tous les modèles ont échoué."
                prog.progress(100, "Terminé !")
                st.session_state.update({"analyses":analyses,"synthesis":synthesis})
            prog.empty()
            valid = len(valid_analyses)
            failed = len(selected_models) - valid
            if valid > 0:
                st.success(f"✅ {valid}/{len(selected_models)} analyses réussies — Coût : **0€** 🎉")
            if failed > 0:
                failed_models = [k for k,v in analyses.items() if v.startswith("❌")]
                st.warning(f"⚠️ {failed} modèle(s) ont échoué : {', '.join([m.split('/')[-1].replace(':free','') for m in failed_models])}")
                with st.expander("Voir les erreurs détaillées"):
                    for m, err in [(k,v) for k,v in analyses.items() if v.startswith("❌")]:
                        st.write(f"**{m}** : {err}")
                    st.info("💡 Essayez de changer de modèle dans la barre latérale ou vérifiez votre connexion.")

        if "synthesis" in st.session_state:
            st.markdown('<div class="section-title">🔀 Synthèse Consolidée</div>', unsafe_allow_html=True)
            st.markdown(st.session_state["synthesis"])

        if "analyses" in st.session_state:
            st.markdown('<div class="section-title">📋 Analyses individuelles</div>', unsafe_allow_html=True)
            for slug, text in st.session_state["analyses"].items():
                name = slug.split("/")[-1].replace(":free","")
                icon = "✅" if not text.startswith("❌") else "❌"
                with st.expander(f"{icon} {name}"):
                    st.markdown(text)

        # ─ Bloc 3 : Q&R libre ────────────────────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="section-title">💬 Question libre sur vos données</div>', unsafe_allow_html=True)
        q = st.text_input("Posez votre question :", placeholder="Ex: Quelle région a les meilleures ventes ?")
        qm = st.selectbox("Modèle :", FREE_MODELS, key="qm")
        if q and st.button("Envoyer ↗", key="qbtn"):
            ov = eda.get("overview",{})
            ctx_q = (f"Données : {ov.get('rows','?')} lignes. "
                     f"Numériques : {', '.join(ov.get('numeric_cols',[])[:8])}. "
                     f"Catégorielles : {', '.join(ov.get('categorical_cols',[])[:5])}. "
                     f"Contexte : {context or 'non précisé'}.")
            with st.spinner("Analyse en cours..."):
                rep, used_m = call_with_fallback(
                    [{"role":"user","content":f"{ctx_q}\n\nQuestion : {q}"}],
                    qm, api,
                    system="Tu es DataMind, analyste expert. Réponds précisément avec des chiffres.",
                    max_tokens=1200, temperature=0.2)
            if rep.startswith("❌"):
                st.error(rep)
                st.info("Vérifiez votre clé OpenRouter ou essayez un autre modèle.")
            else:
                used_name = used_m.split("/")[-1].replace(":free","")
                st.caption(f"Réponse générée par : {used_name}")
                st.markdown(f'<div class="ok-card"><strong>Réponse :</strong><br><br>{rep}</div>', unsafe_allow_html=True)

# ╔══════════════════════╗
# ║  ONGLET 6 RAPPORT   ║
# ╚══════════════════════╝

# ╔══════════════════════════════════════╗
# ║  ONGLET 7 — ML AUTOMATIQUE          ║
# ╚══════════════════════════════════════╝
with tabs[6]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        num_cols = eda["overview"]["numeric_cols"]
        cat_cols = eda["overview"]["categorical_cols"]

        st.markdown('''<div class="insight-card">
<strong>🤖 Machine Learning Automatique</strong><br>
Entraîne automatiquement des modèles ML sur vos données : <strong>Régression</strong>,
<strong>Classification</strong> et <strong>Clustering</strong>. Résultats interprétés par l'IA.
</div>''', unsafe_allow_html=True)

        ml_type = st.radio("Type d'analyse ML :",
                            ["📈 Régression (prédire une valeur numérique)",
                             "🎯 Classification (prédire une catégorie)",
                             "🔵 Clustering (segmenter automatiquement)"],
                            horizontal=True)

        from sklearn.model_selection import train_test_split, cross_val_score
        from sklearn.preprocessing import LabelEncoder, StandardScaler
        from sklearn.linear_model import LinearRegression, LogisticRegression
        from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier, GradientBoostingClassifier
        from sklearn.tree import DecisionTreeClassifier
        from sklearn.cluster import KMeans, AgglomerativeClustering
        from sklearn.metrics import (mean_squared_error, r2_score, mean_absolute_error,
                                      accuracy_score, classification_report, confusion_matrix,
                                      silhouette_score)
        import warnings; warnings.filterwarnings("ignore")

        # ── RÉGRESSION ─────────────────────────────────────────────────────
        if "Régression" in ml_type:
            st.markdown('<div class="section-title">📈 Régression automatique</div>', unsafe_allow_html=True)
            if len(num_cols) < 2:
                st.warning("Régression nécessite au moins 2 variables numériques.")
            else:
                r1,r2 = st.columns(2)
                target_reg = r1.selectbox("Variable cible (Y) :", num_cols, key="reg_y")
                features_reg = r2.multiselect("Variables explicatives (X) :",
                    [c for c in num_cols if c != target_reg],
                    default=[c for c in num_cols if c != target_reg][:5], key="reg_x")

                test_size = st.slider("Taille jeu de test (%)", 10, 40, 20, key="reg_ts") / 100

                if features_reg and st.button("🚀 Entraîner les modèles de régression", key="reg_run"):
                    df_ml = df[features_reg + [target_reg]].dropna()
                    X = df_ml[features_reg].values
                    y = df_ml[target_reg].values

                    scaler = StandardScaler()
                    X_scaled = scaler.fit_transform(X)
                    X_train, X_test, y_train, y_test = train_test_split(
                        X_scaled, y, test_size=test_size, random_state=42)

                    models_reg = {
                        "Régression Linéaire": LinearRegression(),
                        "Random Forest": RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1),
                    }

                    results_reg = []
                    best_model = None
                    best_r2 = -999

                    prog = st.progress(0)
                    for i, (name, model) in enumerate(models_reg.items()):
                        prog.progress((i+1)/len(models_reg), f"Entraînement {name}...")
                        model.fit(X_train, y_train)
                        y_pred = model.predict(X_test)
                        r2  = r2_score(y_test, y_pred)
                        mse = mean_squared_error(y_test, y_pred)
                        mae = mean_absolute_error(y_test, y_pred)
                        cv  = cross_val_score(model, X_scaled, y, cv=5, scoring="r2").mean()
                        results_reg.append({"Modèle":name,"R²":round(r2,4),"MAE":round(mae,4),
                                             "RMSE":round(mse**0.5,4),"R² CV (5-fold)":round(cv,4)})
                        if r2 > best_r2:
                            best_r2 = r2; best_model = (name, model)
                    prog.empty()

                    st.session_state["ml_reg_results"] = results_reg
                    st.session_state["ml_reg_best"] = best_model
                    st.session_state["ml_reg_data"] = (X_scaled, y, features_reg, target_reg, scaler)
                    st.success(f"✅ Meilleur modèle : **{best_model[0]}** (R²={best_r2:.4f})")

                if "ml_reg_results" in st.session_state:
                    res_df = pd.DataFrame(st.session_state["ml_reg_results"])
                    st.dataframe(res_df.style.highlight_max(subset=["R²","R² CV (5-fold)"],color="#1a4731")
                                               .highlight_min(subset=["MAE","RMSE"],color="#1a4731"),
                                  use_container_width=True)

                    # Graphique prédit vs réel
                    X_sc, y_all, feats, tgt, scaler = st.session_state["ml_reg_data"]
                    best_name, best_mdl = st.session_state["ml_reg_best"]
                    y_pred_all = best_mdl.predict(X_sc)
                    fig_pv = px.scatter(x=y_all, y=y_pred_all, opacity=0.6,
                                         color_discrete_sequence=["#F77F00"],
                                         labels={"x":"Valeurs réelles","y":"Valeurs prédites"},
                                         title=f"Prédit vs Réel — {best_name}")
                    mn, mx = float(y_all.min()), float(y_all.max())
                    fig_pv.add_trace(go.Scatter(x=[mn,mx],y=[mn,mx],mode="lines",
                                                 name="Parfait",line=dict(color="#3fb950",dash="dash")))
                    fig_pv.update_layout(**DARK, title_font_size=13)
                    st.plotly_chart(fig_pv, use_container_width=True)

                    # Importance des features (Random Forest)
                    if hasattr(best_mdl, "feature_importances_"):
                        imp_df = pd.DataFrame({"Feature":feats,
                                                "Importance":best_mdl.feature_importances_}).sort_values("Importance",ascending=False)
                        fig_imp = px.bar(imp_df, x="Importance", y="Feature", orientation="h",
                                          color="Importance", color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                                          title="Importance des variables (Feature Importance)")
                        fig_imp.update_layout(**DARK, title_font_size=13, showlegend=False, coloraxis_showscale=False)
                        st.plotly_chart(fig_imp, use_container_width=True)

                    # Interprétation IA
                    api = os.getenv("OPENROUTER_API_KEY","")
                    if api and st.button("🤖 Interpréter les résultats ML avec l'IA", key="reg_ai"):
                        r2_val = st.session_state["ml_reg_results"][0]["R²"]
                        prompt = (
                            "Interprète ces résultats de régression ML :\n"
                            f"Variable cible : {tgt}\n"
                            f"Variables explicatives : {', '.join(feats)}\n"
                            f"Résultats : {json.dumps(st.session_state['ml_reg_results'], ensure_ascii=False)}\n"
                            f"Contexte : {context or 'non précisé'}\n\n"
                            "Explique : qualité du modèle, variables importantes, recommandations business. "
                        )

                        with st.spinner("Analyse IA..."):
                            rep, um = call_with_fallback([{"role":"user","content":prompt}],
                                                          selected_models[0] if selected_models else FREE_MODELS[0],
                                                          api, max_tokens=1500)
                        st.markdown(f'<div class="ok-card">{rep}</div>', unsafe_allow_html=True)

        # ── CLASSIFICATION ─────────────────────────────────────────────────
        elif "Classification" in ml_type:
            st.markdown('<div class="section-title">🎯 Classification automatique</div>', unsafe_allow_html=True)
            if not cat_cols:
                st.warning("Classification nécessite au moins une variable catégorielle cible.")
            else:
                c1,c2 = st.columns(2)
                target_clf = c1.selectbox("Variable cible (Y catégorielle) :", cat_cols, key="clf_y")
                features_clf = c2.multiselect("Variables explicatives (X numériques) :",
                    num_cols, default=num_cols[:5], key="clf_x")
                test_sz = st.slider("Taille jeu de test (%)", 10, 40, 20, key="clf_ts") / 100

                if features_clf and st.button("🚀 Entraîner les modèles de classification", key="clf_run"):
                    df_clf = df[features_clf + [target_clf]].dropna()
                    le = LabelEncoder()
                    y_clf = le.fit_transform(df_clf[target_clf].astype(str))
                    X_clf = df_clf[features_clf].values
                    sc = StandardScaler()
                    X_clf_s = sc.fit_transform(X_clf)
                    X_tr, X_te, y_tr, y_te = train_test_split(X_clf_s, y_clf, test_size=test_sz, random_state=42, stratify=y_clf if len(np.unique(y_clf))>1 else None)

                    models_clf = {
                        "Random Forest":      RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1),
                        "Gradient Boosting":  GradientBoostingClassifier(n_estimators=100, random_state=42),
                        "Arbre de décision":  DecisionTreeClassifier(max_depth=6, random_state=42),
                        "Régression Logistique": LogisticRegression(max_iter=500, random_state=42),
                    }

                    results_clf = []
                    best_clf = None; best_acc = 0
                    prog2 = st.progress(0)
                    for i,(name,mdl) in enumerate(models_clf.items()):
                        prog2.progress((i+1)/len(models_clf), f"Entraînement {name}...")
                        mdl.fit(X_tr, y_tr)
                        y_pr = mdl.predict(X_te)
                        acc = accuracy_score(y_te, y_pr)
                        cv_acc = cross_val_score(mdl, X_clf_s, y_clf, cv=5, scoring="accuracy").mean()
                        results_clf.append({"Modèle":name,"Accuracy":round(acc,4),"Accuracy CV":round(cv_acc,4)})
                        if acc > best_acc: best_acc=acc; best_clf=(name,mdl,y_te,y_pr,le)
                    prog2.empty()

                    st.session_state["ml_clf_results"] = results_clf
                    st.session_state["ml_clf_best"] = best_clf
                    st.session_state["ml_clf_feats"] = features_clf
                    st.success(f"✅ Meilleur modèle : **{best_clf[0]}** (Accuracy={best_acc:.2%})")

                if "ml_clf_results" in st.session_state:
                    res_clf = pd.DataFrame(st.session_state["ml_clf_results"])
                    st.dataframe(res_clf.style.highlight_max(color="#1a4731"), use_container_width=True)

                    bn, bm, y_te, y_pr, le = st.session_state["ml_clf_best"]
                    cm = confusion_matrix(y_te, y_pr)
                    fig_cm = px.imshow(cm, text_auto=True, color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                                        x=le.classes_.tolist(), y=le.classes_.tolist(),
                                        title=f"Matrice de confusion — {bn}")
                    fig_cm.update_layout(**DARK, title_font_size=13)
                    st.plotly_chart(fig_cm, use_container_width=True)

                    if hasattr(bm,"feature_importances_"):
                        feats_clf = st.session_state["ml_clf_feats"]
                        imp2 = pd.DataFrame({"Feature":feats_clf,"Importance":bm.feature_importances_}).sort_values("Importance",ascending=False)
                        fig_i2 = px.bar(imp2,x="Importance",y="Feature",orientation="h",
                                         color="Importance",color_continuous_scale=["#0e1a0f","#009A44","#F77F00"],
                                         title="Importance des variables")
                        fig_i2.update_layout(**DARK,title_font_size=13,showlegend=False,coloraxis_showscale=False)
                        st.plotly_chart(fig_i2, use_container_width=True)

        # ── CLUSTERING ─────────────────────────────────────────────────────
        else:
            st.markdown('<div class="section-title">🔵 Clustering automatique</div>', unsafe_allow_html=True)
            if len(num_cols) < 2:
                st.warning("Clustering nécessite au moins 2 variables numériques.")
            else:
                cl1,cl2 = st.columns(2)
                features_cl = cl1.multiselect("Variables pour le clustering :",
                    num_cols, default=num_cols[:min(5,len(num_cols))], key="cl_x")
                k_max = cl2.slider("Nombre max de clusters (K)", 2, 10, 6, key="cl_k")
                color_cl = st.selectbox("Colorier par :", ["—"]+cat_cols, key="cl_color")

                if features_cl and st.button("🚀 Lancer le clustering", key="cl_run"):
                    df_cl = df[features_cl].dropna()
                    sc_cl = StandardScaler()
                    X_cl = sc_cl.fit_transform(df_cl.values)

                    # Méthode du coude
                    inertias = []
                    silhouettes = []
                    K_range = range(2, min(k_max+1, len(df_cl)//5+2))
                    prog3 = st.progress(0)
                    for i,k in enumerate(K_range):
                        prog3.progress((i+1)/len(K_range), f"Test K={k}...")
                        km = KMeans(n_clusters=k, random_state=42, n_init=10)
                        labels = km.fit_predict(X_cl)
                        inertias.append(km.inertia_)
                        silhouettes.append(silhouette_score(X_cl, labels))
                    prog3.empty()

                    best_k = int(list(K_range)[silhouettes.index(max(silhouettes))])
                    km_best = KMeans(n_clusters=best_k, random_state=42, n_init=10)
                    cluster_labels = km_best.fit_predict(X_cl)
                    df_cl_res = df_cl.copy()
                    df_cl_res["Cluster"] = cluster_labels.astype(str)

                    st.session_state["ml_cl_result"] = (df_cl_res, features_cl, best_k, inertias, silhouettes, list(K_range))
                    st.success(f"✅ Meilleur K : **{best_k} clusters** (Silhouette={max(silhouettes):.4f})")

                if "ml_cl_result" in st.session_state:
                    df_cl_res, feats_cl, best_k, inertias, silhs, K_rng = st.session_state["ml_cl_result"]

                    ec1,ec2 = st.columns(2)
                    with ec1:
                        fig_el = go.Figure()
                        fig_el.add_trace(go.Scatter(x=list(K_rng),y=inertias,mode="lines+markers",
                                                     name="Inertie",line=dict(color="#F77F00",width=2),
                                                     marker=dict(size=8, color="#F77F00")))
                        fig_el.update_layout(title="Méthode du coude",**DARK,
                                              xaxis_title="K clusters",yaxis_title="Inertie",title_font_size=12)
                        st.plotly_chart(fig_el, use_container_width=True)
                    with ec2:
                        fig_sh = go.Figure()
                        fig_sh.add_trace(go.Scatter(x=list(K_rng),y=silhs,mode="lines+markers",
                                                     name="Silhouette",line=dict(color="#009A44",width=2),
                                                     marker=dict(size=8, color="#F77F00")))
                        fig_sh.add_vline(x=best_k,line_dash="dash",line_color="#F77F00",
                                          annotation_text=f"K optimal={best_k}")
                        fig_sh.update_layout(title="Score Silhouette",**DARK,
                                              xaxis_title="K clusters",yaxis_title="Silhouette",title_font_size=12)
                        st.plotly_chart(fig_sh, use_container_width=True)

                    if len(feats_cl) >= 2:
                        fig_cl = px.scatter(df_cl_res, x=feats_cl[0], y=feats_cl[1],
                                             color="Cluster", symbol="Cluster",
                                             color_discrete_sequence=CI_COLORS,
                                             title=f"Clusters K-Means (K={best_k})",opacity=0.75)
                        fig_cl.update_layout(**DARK,title_font_size=13)
                        st.plotly_chart(fig_cl, use_container_width=True)

                    st.markdown('<div class="section-title">Profil des clusters</div>', unsafe_allow_html=True)
                    profile = df_cl_res.groupby("Cluster")[feats_cl].mean().round(3)
                    st.dataframe(profile, use_container_width=True)

                    fig_radar = go.Figure()
                    cats_r = feats_cl
                    norm_p = (profile - profile.min()) / (profile.max() - profile.min() + 1e-8)
                    pal = px.colors.qualitative.Plotly
                    for ci in norm_p.index:
                        vals = norm_p.loc[ci].tolist()
                        fig_radar.add_trace(go.Scatterpolar(
                            r=vals+[vals[0]], theta=cats_r+[cats_r[0]],
                            fill="toself", name=f"Cluster {ci}",
                            line=dict(color=pal[int(ci)%len(pal)])))
                    fig_radar.update_layout(
                        polar=dict(bgcolor="#161b22",
                                    radialaxis=dict(gridcolor="#30363d",color="#8b949e"),
                                    angularaxis=dict(gridcolor="#30363d",color="#e6edf3")),
                        paper_bgcolor="#0d1117",font_color="#e6edf3",
                        title=f"Radar des profils — {best_k} clusters",title_font_size=13,
                        legend=dict(bgcolor="#161b22"))
                    st.plotly_chart(fig_radar, use_container_width=True)


# ╔══════════════════════════════════════╗
# ║  ONGLET 8 — EXPORT EXCEL            ║
# ╚══════════════════════════════════════╝
with tabs[7]:
    if "df" not in st.session_state:
        st.info("👈 Chargez vos données d'abord.")
    else:
        df  = st.session_state["df"]
        eda = st.session_state["eda"]
        st.markdown('<div class="section-title">📊 Export Excel complet</div>', unsafe_allow_html=True)
        st.markdown('<div class="insight-card">Génère un fichier <strong>.xlsx</strong> multi-feuilles avec données nettoyées, statistiques, corrélations, analyses IA et résultats ML.</div>', unsafe_allow_html=True)

        inc1,inc2,inc3,inc4 = st.columns(4)
        inc_data  = inc1.checkbox("Données nettoyées", True)
        inc_stats = inc2.checkbox("Statistiques", True)
        inc_corr  = inc3.checkbox("Corrélations", True)
        inc_ia    = inc4.checkbox("Analyses IA", True)

        if st.button("📥 Générer le fichier Excel", type="primary", use_container_width=True):
            try:
                import io as _io
                from openpyxl import Workbook
                from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.chart import BarChart, Reference

                wb = Workbook()
                BLUE_FILL  = PatternFill("solid", fgColor="1F6FEB")
                DARK_FILL  = PatternFill("solid", fgColor="161B22")
                GREEN_FILL = PatternFill("solid", fgColor="238636")
                WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
                BODY_FONT  = Font(name="Calibri", size=10)
                CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
                HEADER_ALT = PatternFill("solid", fgColor="0D1117")

                def style_header(ws, row=1, fill=BLUE_FILL, font=WHITE_FONT):
                    for cell in ws[row]:
                        cell.fill = fill; cell.font = font; cell.alignment = CENTER

                def write_df(ws, df_in, start_row=1):
                    for r_idx, row in enumerate(dataframe_to_rows(df_in, index=True, header=True), start_row):
                        for c_idx, val in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.font = BODY_FONT
                            if r_idx == start_row:
                                cell.fill = BLUE_FILL; cell.font = WHITE_FONT; cell.alignment = CENTER
                    return r_idx

                # ── Feuille 1 : Résumé ────────────────────────────────────
                ws0 = wb.active; ws0.title = "📋 Résumé"
                ws0.column_dimensions["A"].width = 30
                ws0.column_dimensions["B"].width = 40
                headers_sum = [("Métrique","Valeur")]
                ov = eda["overview"]
                rows_sum = [
                    ("Fichier source", st.session_state.get("source","—")),
                    ("Date d'export", datetime.now().strftime("%d/%m/%Y %H:%M")),
                    ("Lignes", f"{ov['rows']:,}"),
                    ("Colonnes", ov["columns"]),
                    ("Mémoire", f"{ov['memory_mb']} Mo"),
                    ("Valeurs manquantes", f"{ov['missing_pct']}%"),
                    ("Variables numériques", len(ov["numeric_cols"])),
                    ("Variables catégorielles", len(ov["categorical_cols"])),
                    ("Anomalies détectées", int(df["anomalie"].sum()) if "anomalie" in df.columns else "N/A"),
                ]
                ws0.append(["Métrique","Valeur"])
                style_header(ws0)
                for r in rows_sum:
                    ws0.append(list(r))

                # ── Feuille 2 : Données ───────────────────────────────────
                if inc_data:
                    ws1 = wb.create_sheet("📄 Données")
                    write_df(ws1, df.head(10000))
                    for col in ws1.columns:
                        ws1.column_dimensions[col[0].column_letter].width = 15

                # ── Feuille 3 : Statistiques ──────────────────────────────
                if inc_stats and "descriptive_stats" in eda:
                    ws2 = wb.create_sheet("📊 Statistiques")
                    stats_df = pd.DataFrame(eda["descriptive_stats"]).round(4)
                    write_df(ws2, stats_df)
                    for col in ws2.columns:
                        ws2.column_dimensions[col[0].column_letter].width = 18

                    # Distributions
                    row_cur = stats_df.shape[0] + 4
                    ws2.cell(row_cur, 1, "ANALYSE DES DISTRIBUTIONS").fill = BLUE_FILL
                    ws2.cell(row_cur, 1).font = WHITE_FONT
                    row_cur += 1
                    ws2.append(["Variable","Asymétrie","Kurtosis","Outliers","Distribution"])
                    style_header(ws2, row_cur)
                    for col_d, d in eda.get("distributions",{}).items():
                        ws2.append([col_d, d["skewness"], d["kurtosis"],
                                     d["outliers_iqr"], "Normale" if d["is_normal"] else "Non-normale"])

                # ── Feuille 4 : Corrélations ──────────────────────────────
                if inc_corr and "correlations" in eda:
                    ws3 = wb.create_sheet("🔗 Corrélations")
                    corr_df = pd.DataFrame(eda["correlations"]["matrix"]).round(3)
                    write_df(ws3, corr_df)
                    # Paires fortes
                    r_start = corr_df.shape[0] + 3
                    ws3.cell(r_start, 1, "PAIRES FORTEMENT CORRÉLÉES").fill = BLUE_FILL
                    ws3.cell(r_start, 1).font = WHITE_FONT
                    ws3.append(["Variable 1","Variable 2","Corrélation r","Force","Direction"])
                    for p in eda["correlations"]["strong_pairs"]:
                        force = "Très forte" if abs(p["r"])>0.8 else "Forte" if abs(p["r"])>0.65 else "Modérée"
                        ws3.append([p["col1"],p["col2"],p["r"],force,"Positive" if p["r"]>0 else "Négative"])
                    for col in ws3.columns:
                        ws3.column_dimensions[col[0].column_letter].width = 20

                # ── Feuille 5 : Analyses IA ───────────────────────────────
                if inc_ia:
                    ws4 = wb.create_sheet("🤖 Analyses IA")
                    ws4.column_dimensions["A"].width = 25
                    ws4.column_dimensions["B"].width = 100
                    row_ia = 1
                    for section, key in [
                        ("SYNTHÈSE CONSOLIDÉE","synthesis"),
                        ("QUESTIONS IA","ai_questions"),
                        ("RÉPONSES IA","ai_answers"),
                    ]:
                        if key in st.session_state:
                            ws4.cell(row_ia, 1, section).fill = BLUE_FILL
                            ws4.cell(row_ia, 1).font = WHITE_FONT
                            row_ia += 1
                            for line in str(st.session_state[key]).split("\n"):
                                if line.strip():
                                    ws4.cell(row_ia, 1, line.strip())
                                    ws4.cell(row_ia, 1).font = BODY_FONT
                                    ws4.cell(row_ia, 1).alignment = Alignment(wrap_text=True)
                                    row_ia += 1
                            row_ia += 1

                    if "analyses" in st.session_state:
                        ws4.cell(row_ia, 1, "ANALYSES INDIVIDUELLES").fill = GREEN_FILL
                        ws4.cell(row_ia, 1).font = WHITE_FONT
                        row_ia += 1
                        for slug, text in st.session_state["analyses"].items():
                            ws4.cell(row_ia, 1, slug).fill = DARK_FILL
                            ws4.cell(row_ia, 1).font = Font(color="58A6FF", bold=True, name="Calibri")
                            row_ia += 1
                            for line in str(text).split("\n"):
                                if line.strip():
                                    ws4.cell(row_ia, 1, line.strip())
                                    ws4.cell(row_ia, 1).font = BODY_FONT
                                    row_ia += 1
                            row_ia += 1

                # ── Feuille 6 : Résultats ML ──────────────────────────────
                for key, title in [("ml_reg_results","📈 ML Régression"),
                                    ("ml_clf_results","🎯 ML Classification")]:
                    if key in st.session_state:
                        ws_ml = wb.create_sheet(title)
                        ml_df = pd.DataFrame(st.session_state[key])
                        write_df(ws_ml, ml_df)
                        for col in ws_ml.columns:
                            ws_ml.column_dimensions[col[0].column_letter].width = 22

                # Sauvegarde
                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button("⬇️ Télécharger le fichier Excel",
                                    data=buf.read(),
                                    file_name=f"DataMind_Export_{ts}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary", use_container_width=True)
                st.success("✅ Fichier Excel généré avec succès !")
                st.info("📋 Feuilles incluses : Résumé · Données · Statistiques · Corrélations · Analyses IA · Résultats ML")

            except ImportError:
                st.error("openpyxl manquant : `pip install openpyxl`")
            except Exception as e:
                import traceback
                st.error(f"Erreur Excel : {e}")
                st.code(traceback.format_exc())


# ╔══════════════════════════════════════╗
# ║  ONGLET 9 — COMPARAISON FICHIERS    ║
# ╚══════════════════════════════════════╝
with tabs[8]:
    st.markdown('<div class="section-title">📁 Comparaison de plusieurs fichiers</div>', unsafe_allow_html=True)
    st.markdown("<div class=\"insight-card\">Comparez jusqu'a 3 jeux de donnees. Statistiques et insights generés automatiquement.</div>", unsafe_allow_html=True)

    comp_files = st.file_uploader("Chargez 2 ou 3 fichiers à comparer",
                                    type=["csv","xlsx","xls","parquet","json"],
                                    accept_multiple_files=True, key="comp_files")

    if comp_files and len(comp_files) >= 2:
        dfs_comp = {}
        for uf in comp_files[:3]:
            try:
                ext = Path(uf.name).suffix.lower()
                if ext == ".csv":
                    raw = uf.read()
                    for enc in ["utf-8","latin-1","windows-1252","iso-8859-1"]:
                        try:
                            sample = raw[:2048].decode(enc,errors="ignore")
                            sep = ";" if sample.count(";")>sample.count(",") else ","
                            df_c = pd.read_csv(io.BytesIO(raw),encoding=enc,sep=sep,on_bad_lines="skip")
                            break
                        except: continue
                elif ext in (".xlsx",".xls"): df_c = pd.read_excel(uf)
                elif ext == ".parquet": df_c = pd.read_parquet(uf)
                elif ext == ".json": df_c = pd.read_json(uf)
                df_c, _ = auto_clean(df_c)
                dfs_comp[uf.name] = df_c
            except Exception as e:
                st.warning(f"Erreur chargement {uf.name} : {e}")

        if len(dfs_comp) >= 2:
            names = list(dfs_comp.keys())

            # ── Tableau comparatif global ─────────────────────────────────
            st.markdown('<div class="section-title">Vue d\'ensemble comparative</div>', unsafe_allow_html=True)
            comp_overview = []
            for name, df_c in dfs_comp.items():
                eda_c = full_eda(df_c)
                ov_c = eda_c["overview"]
                comp_overview.append({
                    "Fichier": name,
                    "Lignes": f"{ov_c['rows']:,}",
                    "Colonnes": ov_c["columns"],
                    "Mémoire (Mo)": ov_c["memory_mb"],
                    "Manquants (%)": ov_c["missing_pct"],
                    "Var. numériques": len(ov_c["numeric_cols"]),
                    "Var. catégorielles": len(ov_c["categorical_cols"]),
                })
            st.dataframe(pd.DataFrame(comp_overview), use_container_width=True)

            # ── Colonnes communes ─────────────────────────────────────────
            common_num = list(set.intersection(*[
                set(df_c.select_dtypes(include=np.number).columns)
                for df_c in dfs_comp.values()
            ]))

            if common_num:
                st.markdown('<div class="section-title">Comparaison statistique — colonnes communes</div>', unsafe_allow_html=True)
                col_comp = st.selectbox("Variable à comparer :", common_num, key="comp_col")

                # Statistiques côte à côte
                stats_comp = []
                for name, df_c in dfs_comp.items():
                    s = df_c[col_comp].describe()
                    stats_comp.append({
                        "Fichier": name,
                        "Moyenne": round(float(s["mean"]),3),
                        "Écart-type": round(float(s["std"]),3),
                        "Min": round(float(s["min"]),3),
                        "Médiane": round(float(s["50%"]),3),
                        "Max": round(float(s["max"]),3),
                    })
                st.dataframe(pd.DataFrame(stats_comp), use_container_width=True)

                # Distributions superposées
                fig_comp = go.Figure()
                pal = px.colors.qualitative.Plotly
                for i,(name,df_c) in enumerate(dfs_comp.items()):
                    fig_comp.add_trace(go.Histogram(
                        x=df_c[col_comp].dropna(), name=name,
                        opacity=0.6, nbinsx=40,
                        marker_color=pal[i%len(pal)]))
                fig_comp.update_layout(
                    barmode="overlay", title=f"Distributions comparées — {col_comp}",
                    **DARK, title_font_size=13,
                    legend=dict(bgcolor="#161b22"))
                st.plotly_chart(fig_comp, use_container_width=True)

                # Boxplots comparatifs
                all_rows = []
                for name, df_c in dfs_comp.items():
                    tmp = df_c[[col_comp]].dropna().copy()
                    tmp["Source"] = name
                    all_rows.append(tmp)
                df_all = pd.concat(all_rows, ignore_index=True)
                fig_bx = px.box(df_all, x="Source", y=col_comp, color="Source",
                                 color_discrete_sequence=pal,
                                 title=f"Boxplots comparatifs — {col_comp}")
                fig_bx.update_layout(**DARK, title_font_size=13, showlegend=False)
                st.plotly_chart(fig_bx, use_container_width=True)

                # Test statistique de comparaison
                if len(dfs_comp) == 2:
                    vals = [df_c[col_comp].dropna().values for df_c in dfs_comp.values()]
                    t_stat, p_val = stats.ttest_ind(vals[0], vals[1])
                    sig = p_val < 0.05
                    cls = "ok-card" if sig else "insight-card"
                    st.markdown(
                        f'<div class="{cls}">🧪 Test t de Student : t={t_stat:.3f} | p={p_val:.4f} | '
                        f'{"✅ Différence significative" if sig else "⚪ Pas de différence significative"} entre les deux fichiers</div>',
                        unsafe_allow_html=True)

            # ── Analyse IA comparative ────────────────────────────────────
            api_comp = os.getenv("OPENROUTER_API_KEY","")
            if api_comp and common_num:
                st.markdown('<div class="section-title">🤖 Analyse comparative IA</div>', unsafe_allow_html=True)
                if st.button("Generer analyse comparative IA", key="comp_ai"):
                    summary_parts = []
                    for n, df_c in dfs_comp.items():
                        means = df_c[common_num[:3]].mean().round(2).to_dict()
                        summary_parts.append(f"Fichier {n}: {df_c.shape[0]:,} lignes, {df_c.shape[1]} colonnes, moyennes: {means}")
                    summary = "\n".join(summary_parts)
                    lang_r = "francais" if language == "fr" else "English"
                    prompt = (
                        f"Compare ces {len(dfs_comp)} jeux de donnees:\n{summary}\n"
                        f"Colonnes communes : {', '.join(common_num[:8])}\n\n"
                        f"Analyse les differences, tendances et recommandations en {lang_r}."
                    )

                    with st.spinner("Analyse comparative..."):
                        rep_comp, um = call_with_fallback(
                            [{"role":"user","content":prompt}],
                            selected_models[0] if selected_models else FREE_MODELS[0],
                            api_comp, max_tokens=1500)
                    st.markdown(f'<div class="ok-card">{rep_comp}</div>', unsafe_allow_html=True)
            else:
                if not common_num:
                    st.info("Aucune colonne numérique commune trouvée entre les fichiers.")
    elif comp_files and len(comp_files) < 2:
        st.info("Chargez au moins 2 fichiers pour la comparaison.")
    else:
        st.info("👆 Chargez vos fichiers ci-dessus pour commencer la comparaison.")


with tabs[9]:
    if "synthesis" not in st.session_state and "ai_answers" not in st.session_state:
        st.info("👈 Lancez les analyses IA dans l'onglet **IA & Questions** d'abord.")
    else:
        st.markdown('<div class="section-title">Générer le Rapport PDF Détaillé</div>', unsafe_allow_html=True)
        r1,r2 = st.columns(2)
        rpt_title = r1.text_input("Titre", value="Rapport d'Analyse DataMind")
        company   = r2.text_input("Organisation", value="DataMind")

        include_questions = st.checkbox("Inclure les questions/réponses IA", value=True)
        include_stats     = st.checkbox("Inclure les statistiques détaillées", value=True)

        if st.button("📄 Générer le PDF complet", type="primary", use_container_width=True):
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
                from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                                 Table, TableStyle, PageBreak, HRFlowable)

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                pdf_path = OUTPUT_DIR / f"DataMind_Rapport_{ts}.pdf"
                W, H = A4

                CPDF = {
                    "blue":  colors.HexColor("#1f6feb"),
                    "dark":  colors.HexColor("#0d1117"),
                    "gray":  colors.HexColor("#161b22"),
                    "light": colors.HexColor("#e6edf3"),
                    "muted": colors.HexColor("#8b949e"),
                    "green": colors.HexColor("#3fb950"),
                    "warn":  colors.HexColor("#e3b341"),
                }

                def sty(name, **kw):
                    return ParagraphStyle(name, **kw)

                S = {
                    "cover_title": sty("ct", fontSize=24, textColor=CPDF["light"], fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=8),
                    "cover_sub":   sty("cs", fontSize=11, textColor=CPDF["muted"], fontName="Helvetica", alignment=TA_CENTER),
                    "h1": sty("h1", fontSize=15, textColor=CPDF["blue"], fontName="Helvetica-Bold", spaceBefore=18, spaceAfter=6),
                    "h2": sty("h2", fontSize=12, textColor=CPDF["light"], fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4),
                    "h3": sty("h3", fontSize=10, textColor=CPDF["green"], fontName="Helvetica-Bold", spaceBefore=6, spaceAfter=3),
                    "body": sty("body", fontSize=9, textColor=colors.HexColor("#c9d1d9"), fontName="Helvetica", leading=14, alignment=TA_JUSTIFY, spaceAfter=5),
                    "bullet": sty("bullet", fontSize=9, textColor=colors.HexColor("#c9d1d9"), fontName="Helvetica", leading=13, leftIndent=14, spaceAfter=3),
                    "caption": sty("cap", fontSize=7.5, textColor=CPDF["muted"], fontName="Helvetica-Oblique", alignment=TA_CENTER),
                    "code": sty("code", fontSize=8, textColor=colors.HexColor("#58a6ff"), fontName="Helvetica", leading=12, leftIndent=10, spaceBefore=2, spaceAfter=2),
                }

                doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                                         rightMargin=2*cm, leftMargin=2*cm,
                                         topMargin=2*cm, bottomMargin=2*cm)
                E = []

                def add_text(text, default_style="body"):
                    for line in text.split("\n"):
                        line = line.strip()
                        if not line:
                            E.append(Spacer(1, 0.15*cm))
                        elif line.startswith("## "):
                            E.append(Paragraph(line[3:], S["h2"]))
                        elif line.startswith("# "):
                            E.append(Paragraph(line[2:], S["h1"]))
                        elif line.startswith(("- ","• ","* ")):
                            E.append(Paragraph(f"• {line[2:]}", S["bullet"]))
                        elif line.startswith(("P1:","P2:","P3:","Q1:","Q2:","Q3:","Q4:","Q5:","Q6:","Q7:","Q8:","Q9:","Q10:")):
                            E.append(Paragraph(line, S["h3"]))
                        elif line.startswith("Hypothèse:"):
                            E.append(Paragraph(line, S["code"]))
                        else:
                            E.append(Paragraph(line, S[default_style]))

                # ── Couverture ──
                cv = Table([[Paragraph(rpt_title, S["cover_title"])],
                             [Paragraph(f"Généré par {company} | DataMind AI Platform v3.0", S["cover_sub"])],
                             [Paragraph(f"Date : {datetime.now().strftime('%d/%m/%Y à %H:%M')}", S["cover_sub"])],
                             [Paragraph(f"Source : {st.session_state.get('source','—')}", S["cover_sub"])]],
                            colWidths=[W-4*cm])
                cv.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#0e1a0f")),
                                         ("TOPPADDING",(0,0),(-1,0),55),("BOTTOMPADDING",(0,-1),(-1,-1),55),
                                         ("LEFTPADDING",(0,0),(-1,-1),30),("RIGHTPADDING",(0,0),(-1,-1),30)]))
                E.append(cv)
                E.append(Spacer(1,1*cm))

                # Infos
                ov = st.session_state["eda"].get("overview",{})
                info = [
                    ["📊 Données","f{ov.get('rows','?'):,} lignes × {ov.get('columns','?')} colonnes"],
                    ["🤖 Modèles IA",", ".join(list(st.session_state.get("analyses",{}).keys())[:3]) or "—"],
                    ["🌍 Langue","Français" if language=="fr" else "English"],
                    ["📝 Contexte", context or "Non précisé"],
                ]
                info[0][1] = f"{ov.get('rows','?'):,} lignes × {ov.get('columns','?')} colonnes"
                it = Table(info, colWidths=[4*cm, 12*cm])
                it.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(0,-1),CPDF["gray"]),
                    ("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),9),
                    ("TEXTCOLOR",(0,0),(0,-1),colors.HexColor("#58a6ff")),
                    ("TEXTCOLOR",(1,0),(1,-1),colors.HexColor("#c9d1d9")),
                    ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#30363d")),
                    ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                    ("LEFTPADDING",(0,0),(-1,-1),10),
                ]))
                E.append(it)
                E.append(PageBreak())

                # ── Statistiques ──
                if include_stats and "descriptive_stats" in st.session_state["eda"]:
                    E.append(Paragraph("Statistiques Descriptives", S["h1"]))
                    E.append(HRFlowable(width="100%", thickness=1.5, color=CPDF["blue"]))
                    E.append(Spacer(1,0.3*cm))
                    ds = st.session_state["eda"]["descriptive_stats"]
                    cols_pdf = list(ds.keys())[:8]
                    header = ["Stat"] + [c[:10] for c in cols_pdf]
                    rows_s = [header]
                    for met in ["count","mean","std","min","25%","50%","75%","max"]:
                        row = [met]
                        for c in cols_pdf:
                            v = ds[c].get(met,"—")
                            row.append(f"{float(v):.3g}" if isinstance(v,(int,float)) else "—")
                        rows_s.append(row)
                    cw = (W-4*cm)/(len(cols_pdf)+1)
                    st_tbl = Table(rows_s, colWidths=[cw]*(len(cols_pdf)+1))
                    st_tbl.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(-1,0),CPDF["blue"]),
                        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                        ("FONTSIZE",(0,0),(-1,-1),7.5),
                        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f0f4f8")]),
                        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#c5cedf")),
                        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                    ]))
                    E.append(st_tbl)
                    E.append(Spacer(1,0.5*cm))

                    # Corrélations
                    strong = st.session_state["eda"].get("correlations",{}).get("strong_pairs",[])
                    if strong:
                        E.append(Paragraph("Corrélations significatives", S["h2"]))
                        for p in strong[:8]:
                            dir_ = "positive" if p["r"]>0 else "négative"
                            force = "très forte" if abs(p["r"])>0.8 else "forte" if abs(p["r"])>0.65 else "modérée"
                            E.append(Paragraph(
                                f"• {p['col1']} ↔ {p['col2']} : r={p['r']} — Corrélation {force} {dir_}",
                                S["bullet"]))
                    E.append(PageBreak())

                # ── Questions/Réponses IA ──
                if include_questions and "ai_questions" in st.session_state:
                    E.append(Paragraph("Questions Analytiques Générées par l'IA", S["h1"]))
                    E.append(HRFlowable(width="100%",thickness=1.5,color=CPDF["green"]))
                    E.append(Spacer(1,0.3*cm))
                    E.append(Paragraph("L'IA a auto-généré ces questions après analyse de vos données :", S["body"]))
                    E.append(Spacer(1,0.2*cm))
                    add_text(st.session_state["ai_questions"])
                    E.append(Spacer(1,0.5*cm))
                    if "ai_answers" in st.session_state:
                        E.append(Paragraph("Réponses basées sur les données réelles", S["h2"]))
                        add_text(st.session_state["ai_answers"])
                    E.append(PageBreak())

                # ── Synthèse ──
                if "synthesis" in st.session_state:
                    E.append(Paragraph("Synthèse Consolidée Multi-Modèles", S["h1"]))
                    E.append(HRFlowable(width="100%",thickness=1.5,color=CPDF["blue"]))
                    E.append(Spacer(1,0.3*cm))
                    add_text(st.session_state["synthesis"])
                    E.append(PageBreak())

                # ── Analyses individuelles ──
                if "analyses" in st.session_state:
                    E.append(Paragraph("Analyses Détaillées par Modèle IA", S["h1"]))
                    E.append(HRFlowable(width="100%",thickness=1.5,color=CPDF["blue"]))
                    E.append(Spacer(1,0.3*cm))
                    for slug, text in st.session_state["analyses"].items():
                        if text.startswith("❌"): continue
                        E.append(Paragraph(slug, S["code"]))
                        E.append(HRFlowable(width="100%",thickness=0.4,color=CPDF["gray"]))
                        E.append(Spacer(1,0.2*cm))
                        add_text(text)
                        E.append(Spacer(1,0.8*cm))

                # ── Footer ──
                E.append(HRFlowable(width="100%",thickness=0.5,color=CPDF["muted"]))
                E.append(Spacer(1,0.2*cm))
                E.append(Paragraph(
                    f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | "
                    f"{company} | DataMind AI Platform v3.0 | OpenRouter Free Models",
                    S["caption"]))

                doc.build(E)
                with open(pdf_path,"rb") as f:
                    st.download_button("⬇️ Télécharger le rapport PDF", data=f.read(),
                                        file_name=pdf_path.name, mime="application/pdf",
                                        type="primary", use_container_width=True)
                st.success(f"✅ **{pdf_path.name}** généré avec succès !")

            except ImportError:
                st.error("ReportLab manquant : `pip install reportlab`")
            except Exception as e:
                import traceback
                st.error(f"Erreur PDF : {e}")
                st.code(traceback.format_exc())


# ╔══════════════════════════════════════╗
# ║  ONGLET 11 — ENVOI PAR EMAIL        ║
# ╚══════════════════════════════════════╝
with tabs[10]:
    st.markdown('<div class="section-title">📧 Envoi automatique du rapport par email</div>', unsafe_allow_html=True)
    st.markdown("""<div class="insight-card">
Envoyez le rapport PDF et les analyses IA directement par email.
Supporte <strong>Gmail</strong>, <strong>Outlook</strong>, <strong>Yahoo</strong> et tout serveur SMTP.
Compatible avec les comptes professionnels (entreprise, université).
</div>""", unsafe_allow_html=True)

    # ── Configuration SMTP ────────────────────────────────────────────────
    st.markdown('<div class="section-title">1. Configuration email</div>', unsafe_allow_html=True)

    provider = st.selectbox("Fournisseur email :", [
        "Gmail (smtp.gmail.com)",
        "Outlook / Hotmail (smtp.live.com)",
        "Yahoo (smtp.mail.yahoo.com)",
        "Orange / Wanadoo (smtp.orange.fr)",
        "Serveur SMTP personnalisé",
    ], key="email_provider")

    SMTP_PRESETS = {
        "Gmail (smtp.gmail.com)":           ("smtp.gmail.com", 587),
        "Outlook / Hotmail (smtp.live.com)":("smtp-mail.outlook.com", 587),
        "Yahoo (smtp.mail.yahoo.com)":      ("smtp.mail.yahoo.com", 587),
        "Orange / Wanadoo (smtp.orange.fr)":("smtp.orange.fr", 587),
        "Serveur SMTP personnalisé":        ("", 587),
    }
    smtp_host_default, smtp_port_default = SMTP_PRESETS[provider]

    ec1, ec2 = st.columns(2)
    smtp_host = ec1.text_input("Serveur SMTP :", value=smtp_host_default, key="smtp_host")
    smtp_port = ec2.number_input("Port :", value=smtp_port_default, min_value=1, max_value=65535, key="smtp_port")

    ea1, ea2 = st.columns(2)
    email_from = ea1.text_input("Votre email (expéditeur) :", placeholder="votre@gmail.com", key="email_from")
    email_pass  = ea2.text_input("Mot de passe / Mot de passe d'application :", type="password", key="email_pass",
                                   help="Gmail : utilisez un 'Mot de passe d'application' (2FA requis)")

    if provider == "Gmail (smtp.gmail.com)":
        st.info("ℹ️ **Gmail** : activez la 2FA puis créez un 'Mot de passe d\'application' sur myaccount.google.com → Sécurité → Mots de passe des applications")

    # ── Destinataires ─────────────────────────────────────────────────────
    st.markdown('<div class="section-title">2. Destinataires et contenu</div>', unsafe_allow_html=True)

    email_to_raw = st.text_input("Destinataires (séparés par des virgules) :",
                                   placeholder="colleague@example.com, manager@company.com", key="email_to")
    email_subject = st.text_input("Objet de l'email :",
                                    value=f"Rapport DataMind — {datetime.now().strftime('%d/%m/%Y')}",
                                    key="email_subject")
    email_body = st.text_area("Message personnalisé :", height=120, key="email_body",
                                value="Bonjour,\n\nVeuillez trouver ci-joint le rapport d'analyse de données généré par DataMind AI Platform.\n\nCordialement,\nDataMind")

    # Options pièces jointes
    st.markdown('<div class="section-title">3. Pièces jointes</div>', unsafe_allow_html=True)
    at1, at2, at3 = st.columns(3)
    attach_pdf   = at1.checkbox("Rapport PDF", value=True, key="att_pdf")
    attach_excel = at2.checkbox("Export Excel", value=False, key="att_excel")
    attach_synth = at3.checkbox("Synthèse IA (texte)", value=True, key="att_synth")

    # ── Test de connexion ─────────────────────────────────────────────────
    st.markdown('<div class="section-title">4. Envoi</div>', unsafe_allow_html=True)

    col_test, col_send = st.columns(2)

    with col_test:
        if st.button("🔌 Tester la connexion SMTP", use_container_width=True):
            if not email_from or not email_pass:
                st.error("Remplissez l'email et le mot de passe.")
            else:
                try:
                    import smtplib
                    with smtplib.SMTP(smtp_host, int(smtp_port), timeout=10) as server:
                        server.ehlo()
                        server.starttls()
                        server.login(email_from, email_pass)
                    st.success("✅ Connexion SMTP réussie !")
                except smtplib.SMTPAuthenticationError:
                    st.error("❌ Authentification échouée. Vérifiez email/mot de passe.")
                except smtplib.SMTPConnectError:
                    st.error(f"❌ Impossible de se connecter à {smtp_host}:{smtp_port}")
                except Exception as e:
                    st.error(f"❌ Erreur : {e}")

    with col_send:
        send_btn = st.button("📤 Envoyer le rapport", type="primary", use_container_width=True)

    if send_btn:
        if not email_from or not email_pass:
            st.error("Configurez votre email et mot de passe.")
        elif not email_to_raw.strip():
            st.error("Ajoutez au moins un destinataire.")
        else:
            try:
                import smtplib
                from email.mime.multipart import MIMEMultipart
                from email.mime.text    import MIMEText
                from email.mime.base    import MIMEBase
                from email              import encoders
                import io as _io

                recipients = [e.strip() for e in email_to_raw.split(",") if e.strip()]
                msg = MIMEMultipart()
                msg["From"]    = email_from
                msg["To"]      = ", ".join(recipients)
                msg["Subject"] = email_subject

                # Corps HTML
                synthesis_preview = ""
                if "synthesis" in st.session_state:
                    preview = str(st.session_state["synthesis"])[:1500]
                    synthesis_preview = f"<hr><h3>Extrait de la synthese IA</h3><pre style='font-size:12px'>{preview}...</pre>"

                html_body = f"""
<html><body style="font-family:Arial,sans-serif;color:#2a2e3b;">
<div style="background:#1a3a6b;padding:20px;border-radius:8px;margin-bottom:20px;">
  <h1 style="color:white;margin:0;">DataMind AI Platform</h1>
  <p style="color:#c8d8f5;margin:5px 0 0;">Rapport d'analyse genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}</p>
</div>
<div style="padding:20px;">
  <p>{email_body.replace(chr(10),'<br>')}</p>
  <table style="border-collapse:collapse;width:100%;margin:20px 0;">
    <tr style="background:#1a3a6b;color:white;">
      <th style="padding:10px;text-align:left;">Metrique</th>
      <th style="padding:10px;text-align:left;">Valeur</th>
    </tr>
    <tr style="background:#f4f6fb;">
      <td style="padding:8px;">Source</td>
      <td style="padding:8px;">{st.session_state.get("source","—")}</td>
    </tr>
    <tr>
      <td style="padding:8px;">Lignes analysees</td>
      <td style="padding:8px;">{st.session_state["eda"]["overview"].get("rows","—"):,} lignes</td>
    </tr>
    <tr style="background:#f4f6fb;">
      <td style="padding:8px;">Colonnes</td>
      <td style="padding:8px;">{st.session_state["eda"]["overview"].get("columns","—")}</td>
    </tr>
    <tr>
      <td style="padding:8px;">Modeles IA utilises</td>
      <td style="padding:8px;">{", ".join(list(st.session_state.get("analyses",{}).keys())[:3])}</td>
    </tr>
  </table>
  {synthesis_preview}
  <p style="color:#8b949e;font-size:12px;margin-top:30px;">
    Genere par DataMind AI Platform | OpenRouter Free Models
  </p>
</div>
</body></html>"""

                msg.attach(MIMEText(html_body, "html", "utf-8"))

                # Pièce jointe : Synthèse texte
                if attach_synth and "synthesis" in st.session_state:
                    synth_bytes = st.session_state["synthesis"].encode("utf-8")
                    part = MIMEBase("application","octet-stream")
                    part.set_payload(synth_bytes)
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition","attachment",
                                     filename=f"DataMind_Synthese_{datetime.now().strftime('%Y%m%d')}.txt")
                    msg.attach(part)

                # Pièce jointe : Rapport PDF
                if attach_pdf:
                    pdf_files = sorted(OUTPUT_DIR.glob("DataMind_Rapport_*.pdf"), reverse=True)
                    if pdf_files:
                        with open(pdf_files[0],"rb") as f_pdf:
                            part = MIMEBase("application","octet-stream")
                            part.set_payload(f_pdf.read())
                            encoders.encode_base64(part)
                            part.add_header("Content-Disposition","attachment",
                                             filename=pdf_files[0].name)
                            msg.attach(part)
                    else:
                        st.warning("Aucun PDF trouvé. Générez d'abord le rapport dans l'onglet Rapport.")

                # Pièce jointe : Excel
                if attach_excel:
                    xlsx_files = sorted(OUTPUT_DIR.glob("DataMind_Export_*.xlsx"), reverse=True)
                    if xlsx_files:
                        with open(xlsx_files[0],"rb") as f_xl:
                            part = MIMEBase("application","octet-stream")
                            part.set_payload(f_xl.read())
                            encoders.encode_base64(part)
                            part.add_header("Content-Disposition","attachment",
                                             filename=xlsx_files[0].name)
                            msg.attach(part)

                # Envoi
                with st.spinner(f"Envoi vers {len(recipients)} destinataire(s)..."):
                    with smtplib.SMTP(smtp_host, int(smtp_port), timeout=30) as server:
                        server.ehlo()
                        server.starttls()
                        server.login(email_from, email_pass)
                        server.sendmail(email_from, recipients, msg.as_string())

                st.success(f"✅ Email envoyé avec succès à : {', '.join(recipients)}")
                st.balloons()

            except smtplib.SMTPAuthenticationError:
                st.error("❌ Authentification échouée. Pour Gmail : utilisez un mot de passe d'application.")
            except Exception as e:
                st.error(f"❌ Erreur d'envoi : {e}")
                import traceback; st.code(traceback.format_exc())


# ╔══════════════════════════════════════╗
# ║  ONGLET 12 — BASE DE DONNÉES        ║
# ╚══════════════════════════════════════╝
with tabs[11]:
    st.markdown('<div class="section-title">🗄️ Base de données — Sauvegarde des analyses</div>', unsafe_allow_html=True)
    st.markdown("""<div class="insight-card">
Sauvegardez et retrouvez toutes vos analyses. Fonctionne en <strong>mode local (SQLite)</strong> sans configuration,
ou connectez une base <strong>PostgreSQL / MySQL</strong> distante pour partager entre équipes.
</div>""", unsafe_allow_html=True)

    import sqlite3
    import hashlib

    DB_PATH = OUTPUT_DIR / "datamind_analyses.db"

    def init_db():
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS analyses (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at  TEXT NOT NULL,
                source_name TEXT,
                rows        INTEGER,
                columns     INTEGER,
                context     TEXT,
                language    TEXT,
                models_used TEXT,
                synthesis   TEXT,
                questions   TEXT,
                answers     TEXT,
                eda_json    TEXT,
                tags        TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ml_results (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                analysis_id INTEGER,
                ml_type     TEXT,
                target      TEXT,
                features    TEXT,
                results_json TEXT,
                created_at  TEXT,
                FOREIGN KEY (analysis_id) REFERENCES analyses(id)
            )
        """)
        conn.commit()
        conn.close()

    def save_analysis(source, eda, synthesis, questions, answers, models_used, context, language, tags=""):
        init_db()
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        ov = eda.get("overview", {})
        c.execute("""
            INSERT INTO analyses
            (created_at, source_name, rows, columns, context, language,
             models_used, synthesis, questions, answers, eda_json, tags)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            source, ov.get("rows", 0), ov.get("columns", 0),
            context, language,
            ", ".join(models_used) if isinstance(models_used, list) else str(models_used),
            synthesis or "", questions or "", answers or "",
            json.dumps(eda, ensure_ascii=False)[:50000],
            tags
        ))
        analysis_id = c.lastrowid
        conn.commit()
        conn.close()
        return analysis_id

    def load_analyses():
        init_db()
        conn = sqlite3.connect(str(DB_PATH))
        df_db = pd.read_sql_query(
            "SELECT id, created_at, source_name, rows, columns, context, models_used, tags FROM analyses ORDER BY created_at DESC",
            conn)
        conn.close()
        return df_db

    def load_analysis_detail(analysis_id):
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("SELECT * FROM analyses WHERE id=?", (analysis_id,))
        row = c.fetchone()
        cols = [d[0] for d in c.description]
        conn.close()
        return dict(zip(cols, row)) if row else None

    def delete_analysis(analysis_id):
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        c.execute("DELETE FROM analyses WHERE id=?", (analysis_id,))
        c.execute("DELETE FROM ml_results WHERE analysis_id=?", (analysis_id,))
        conn.commit()
        conn.close()

    init_db()

    db_tabs = st.tabs(["💾 Sauvegarder", "📂 Historique", "🔍 Consulter", "📤 Exporter DB"])

    # ── Sous-onglet 1 : Sauvegarder ──────────────────────────────────────
    with db_tabs[0]:
        st.markdown('<div class="section-title">Sauvegarder l\'analyse actuelle</div>', unsafe_allow_html=True)

        if "eda" not in st.session_state:
            st.info("👈 Chargez et analysez des données d'abord.")
        else:
            ov_db = st.session_state["eda"].get("overview", {})
            db1, db2 = st.columns(2)
            db1.metric("Source", st.session_state.get("source", "—"))
            db2.metric("Données", f"{ov_db.get('rows',0):,} lignes × {ov_db.get('columns',0)} colonnes")

            tags_input = st.text_input("Tags (optionnel) :", placeholder="ventes, 2025, Abidjan, marketing")
            note_input = st.text_area("Note / commentaire :", placeholder="Analyse mensuelle des ventes Q1...", height=80)

            has_synthesis = "synthesis" in st.session_state
            has_questions = "ai_questions" in st.session_state

            st.markdown(f"""
<div class="{'ok-card' if has_synthesis else 'warn-card'}">
{'✅' if has_synthesis else '⚠️'} Synthèse IA : {'disponible' if has_synthesis else 'non générée'}<br>
{'✅' if has_questions else '⚠️'} Questions/Réponses : {'disponibles' if has_questions else 'non générées'}
</div>""", unsafe_allow_html=True)

            if st.button("💾 Sauvegarder cette analyse", type="primary", use_container_width=True):
                with st.spinner("Sauvegarde en cours..."):
                    aid = save_analysis(
                        source    = st.session_state.get("source", "inconnu"),
                        eda       = st.session_state["eda"],
                        synthesis = st.session_state.get("synthesis", ""),
                        questions = st.session_state.get("ai_questions", ""),
                        answers   = st.session_state.get("ai_answers", ""),
                        models_used = list(st.session_state.get("analyses", {}).keys()),
                        context   = context,
                        language  = language,
                        tags      = f"{tags_input} {note_input}".strip(),
                    )
                st.success(f"✅ Analyse sauvegardée avec l'ID **#{aid}** dans `{DB_PATH.name}`")
                st.info("Retrouvez-la dans l'onglet **Historique**.")

    # ── Sous-onglet 2 : Historique ────────────────────────────────────────
    with db_tabs[1]:
        st.markdown('<div class="section-title">Historique des analyses sauvegardées</div>', unsafe_allow_html=True)

        df_hist = load_analyses()

        if df_hist.empty:
            st.info("Aucune analyse sauvegardée. Utilisez l'onglet **Sauvegarder**.")
        else:
            # Métriques rapides
            h1,h2,h3 = st.columns(3)
            h1.markdown(f'<div class="metric-box"><div class="metric-val">{len(df_hist)}</div><div class="metric-lbl">Analyses</div></div>', unsafe_allow_html=True)
            h2.markdown(f'<div class="metric-box"><div class="metric-val">{df_hist["rows"].sum():,}</div><div class="metric-lbl">Lignes totales</div></div>', unsafe_allow_html=True)
            h3.markdown(f'<div class="metric-box"><div class="metric-val">{df_hist["source_name"].nunique()}</div><div class="metric-lbl">Sources uniques</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Filtres
            f1, f2 = st.columns(2)
            search = f1.text_input("Rechercher :", placeholder="nom fichier, tag...")
            sources = ["Tous"] + df_hist["source_name"].dropna().unique().tolist()
            src_filter = f2.selectbox("Filtrer par source :", sources)

            df_show = df_hist.copy()
            if search:
                mask = (df_show["source_name"].str.contains(search, case=False, na=False) |
                        df_show["tags"].str.contains(search, case=False, na=False))
                df_show = df_show[mask]
            if src_filter != "Tous":
                df_show = df_show[df_show["source_name"] == src_filter]

            df_show.columns = ["ID","Date","Source","Lignes","Colonnes","Contexte","Modèles","Tags"]
            st.dataframe(df_show, use_container_width=True, height=350)

            # Suppression
            st.markdown("---")
            del_id = st.number_input("ID à supprimer :", min_value=1, step=1, key="del_id")
            if st.button("🗑️ Supprimer cette analyse", key="del_btn"):
                delete_analysis(int(del_id))
                st.success(f"✅ Analyse #{del_id} supprimée.")
                st.rerun()

    # ── Sous-onglet 3 : Consulter ─────────────────────────────────────────
    with db_tabs[2]:
        st.markdown('<div class="section-title">Consulter une analyse sauvegardée</div>', unsafe_allow_html=True)

        df_list = load_analyses()
        if df_list.empty:
            st.info("Aucune analyse disponible.")
        else:
            options = {f"#{row['id']} — {row['source_name']} ({row['created_at']})": row['id']
                       for _, row in df_list.iterrows()}
            selected_label = st.selectbox("Choisir une analyse :", list(options.keys()))
            selected_id = options[selected_label]

            if st.button("📂 Charger cette analyse", type="primary"):
                detail = load_analysis_detail(selected_id)
                if detail:
                    st.session_state["db_loaded"] = detail
                    st.success(f"✅ Analyse #{selected_id} chargée !")

            if "db_loaded" in st.session_state:
                d = st.session_state["db_loaded"]
                st.markdown(f"""
<div class="insight-card">
<strong>📋 Analyse #{d['id']}</strong><br>
📅 Date : {d['created_at']}<br>
📁 Source : {d['source_name']}<br>
📊 Données : {d['rows']:,} lignes × {d['columns']} colonnes<br>
🌍 Langue : {d['language']}<br>
🤖 Modèles : {d['models_used']}<br>
🏷️ Tags : {d['tags'] or '—'}
</div>""", unsafe_allow_html=True)

                if d.get("synthesis"):
                    with st.expander("🔀 Synthèse consolidée", expanded=True):
                        st.markdown(d["synthesis"])

                if d.get("questions"):
                    with st.expander("❓ Questions IA"):
                        st.markdown(d["questions"])

                if d.get("answers"):
                    with st.expander("💡 Réponses IA"):
                        st.markdown(d["answers"])

                # Restaurer dans la session
                if st.button("♻️ Restaurer cette analyse comme active"):
                    try:
                        eda_restored = json.loads(d["eda_json"])
                        st.session_state["eda"]       = eda_restored
                        st.session_state["synthesis"] = d["synthesis"]
                        st.session_state["source"]    = d["source_name"]
                        if d.get("questions"):
                            st.session_state["ai_questions"] = d["questions"]
                        if d.get("answers"):
                            st.session_state["ai_answers"]   = d["answers"]
                        st.success("✅ Analyse restaurée ! Vous pouvez regénérer des graphiques et rapports.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur restauration : {e}")

    # ── Sous-onglet 4 : Exporter DB ───────────────────────────────────────
    with db_tabs[3]:
        st.markdown('<div class="section-title">Exporter et sauvegarder la base de données</div>', unsafe_allow_html=True)

        df_export = load_analyses()

        if df_export.empty:
            st.info("Aucune donnée à exporter.")
        else:
            # Export CSV
            csv_data = df_export.to_csv(index=False, encoding="utf-8")
            st.download_button(
                "⬇️ Exporter l'historique en CSV",
                data=csv_data.encode("utf-8"),
                file_name=f"DataMind_Historique_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

            # Export fichier SQLite complet
            if DB_PATH.exists():
                with open(DB_PATH, "rb") as f_db:
                    st.download_button(
                        "⬇️ Télécharger la base SQLite complète (.db)",
                        data=f_db.read(),
                        file_name=f"DataMind_DB_{datetime.now().strftime('%Y%m%d')}.db",
                        mime="application/octet-stream",
                        use_container_width=True,
                    )

            # Statistiques globales
            st.markdown('<div class="section-title">Statistiques de la base</div>', unsafe_allow_html=True)
            st.markdown(f"""
<div class="insight-card">
📊 <strong>{len(df_export)}</strong> analyses sauvegardées<br>
📁 <strong>{df_export['source_name'].nunique()}</strong> sources différentes<br>
📅 Première analyse : <strong>{df_export['created_at'].min()}</strong><br>
📅 Dernière analyse : <strong>{df_export['created_at'].max()}</strong><br>
💾 Taille de la base : <strong>{DB_PATH.stat().st_size / 1024:.1f} Ko</strong>
</div>""", unsafe_allow_html=True)

            # Graphique évolution dans le temps
            if len(df_export) > 1:
                df_export["created_at"] = pd.to_datetime(df_export["created_at"])
                df_export["date"] = df_export["created_at"].dt.date
                daily = df_export.groupby("date").size().reset_index(name="analyses")
                fig_db = px.bar(daily, x="date", y="analyses",
                                 title="Analyses par date",
                                 color_discrete_sequence=["#F77F00"])
                fig_db.update_layout(**DARK, title_font_size=12)
                st.plotly_chart(fig_db, use_container_width=True)
